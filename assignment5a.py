"""
APM Project 2026 - Assignment 5a
"""

import gurobipy as gp
from gurobipy import GRB
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from input_data import (
    PARTS, END_PRODUCT, T, BOM, LEAD_TIME, MIN_LOT,
    INIT_INV, SETUP_COST, HOLDING_COST, DEMAND_FORECAST
)

def get_parents(part):
    parents = {}
    for parent, children in BOM.items():
        if part in children:
            parents[parent] = children[part]
    return parents

def clean_num(val, tol=1e-6):
    v = float(val)
    return 0.0 if abs(v) < tol else round(v, 6)

CAP_X_BASE     = 800
CAP_X_MAX_EXP  = 200
COST_EXP_X     = 10
CAP_X_OT_MAX   = 300
COST_OT_X      = 2

CAP_Y_BASE     = 60 * 24 * 7 - 80
CAP_Y_MAX_PCT  = 40
COST_EXP_Y_PCT = 1_500
CAP_Y_OT_MAX   = 38
COST_OT_Y      = 120

PROC_X = {END_PRODUCT: 1}
PROC_Y = {"B1401": 3, "B2302": 2}

model = gp.Model("APM_Assignment5a")
periods = range(1, T + 1)
parts   = PARTS

x      = model.addVars(parts, periods, name="x",  vtype=GRB.INTEGER, lb=0)
y      = model.addVars(parts, periods, name="y",  vtype=GRB.BINARY)
I      = model.addVars(parts, periods, name="I",  vtype=GRB.INTEGER, lb=0)
dx     = model.addVar(name="dx",     vtype=GRB.INTEGER, lb=0, ub=CAP_X_MAX_EXP)
dy_pct = model.addVar(name="dy_pct", lb=0, ub=CAP_Y_MAX_PCT)
ot_x   = model.addVars(periods, name="ot_x", vtype=GRB.INTEGER, lb=0, ub=CAP_X_OT_MAX)
ot_y   = model.addVars(periods, name="ot_y", lb=0, ub=CAP_Y_OT_MAX)

BIG_M = {i: sum(DEMAND_FORECAST) * 25 for i in parts}

model.setObjective(
    gp.quicksum(SETUP_COST[i]*y[i,t] + HOLDING_COST[i]*I[i,t]
                for i in parts for t in periods)
    + COST_EXP_X*dx + COST_EXP_Y_PCT*dy_pct
    + gp.quicksum(COST_OT_X*ot_x[t] + COST_OT_Y*ot_y[t] for t in periods),
    GRB.MINIMIZE)

for i in parts:
    for t in periods:
        inv_prev = INIT_INV[i] if t == 1 else I[i, t-1]
        op = t - LEAD_TIME[i]
        rec = x[i, op] if op >= 1 else 0
        int_d = gp.quicksum(BOM[p][i]*x[p,t] for p in get_parents(i))
        ext_d = DEMAND_FORECAST[t-1] if i == END_PRODUCT else 0
        model.addConstr(I[i,t] == inv_prev + rec - int_d - ext_d,
                        name=f"inv_balance_{i}_{t}")
        model.addConstr(x[i,t] >= MIN_LOT[i]*y[i,t], name=f"min_lot_{i}_{t}")
        model.addConstr(x[i,t] <= BIG_M[i]*y[i,t],   name=f"bigM_{i}_{t}")

for t in periods:
    model.addConstr(
        gp.quicksum(PROC_X[i]*x[i,t] for i in PROC_X if i in parts)
        <= CAP_X_BASE + dx + ot_x[t], name=f"cap_X_{t}")
    model.addConstr(
        gp.quicksum(PROC_Y[i]*x[i,t] for i in PROC_Y if i in parts)
        <= CAP_Y_BASE + (CAP_Y_BASE/100.0)*dy_pct + 60.0*ot_y[t],
        name=f"cap_Y_{t}")

model.optimize()

if model.status == GRB.OPTIMAL:
    dx_val     = clean_num(dx.X)
    dy_pct_val = clean_num(dy_pct.X)
    invest_X   = COST_EXP_X     * dx_val
    invest_Y   = COST_EXP_Y_PCT * dy_pct_val
    cap_x_new  = CAP_X_BASE + dx_val
    cap_y_new  = CAP_Y_BASE * (1 + dy_pct_val/100)

    total_setup   = sum(SETUP_COST[i]   * y[i,t].X for i in parts for t in periods)
    total_holding = sum(HOLDING_COST[i] * I[i,t].X for i in parts for t in periods)
    total_ot_x    = sum(COST_OT_X * ot_x[t].X for t in periods)
    total_ot_y    = sum(COST_OT_Y * ot_y[t].X for t in periods)

    OUTPUT_FILE = "OUTPUT.xlsx"
    SHEET_NAME  = "Output_5a"

    if os.path.exists(OUTPUT_FILE):
        wb = load_workbook(OUTPUT_FILE)
        if SHEET_NAME in wb.sheetnames:
            del wb[SHEET_NAME]
        ws = wb.create_sheet(SHEET_NAME)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME

    # ── Styles ────────────────────────────────────────────────────────────
    NO_FILL    = PatternFill(fill_type=None)
    BLACK_FILL = PatternFill("solid", start_color="000000", end_color="000000")
    none_border = Border()
    bot_medium  = Border(bottom=Side(style="medium", color="000000"))
    bot_thin    = Border(bottom=Side(style="thin",   color="CCCCCC"))

    def plain(cell, val, bold=False, align="left", size=9, color="000000", fmt=None, border=None):
        cell.value = val
        cell.font  = Font(name="Calibri", bold=bold, size=size, color=color)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.fill  = NO_FILL
        if border: cell.border = border
        if fmt:    cell.number_format = fmt

    def hdr(cell, val):
        cell.value = val
        cell.font  = Font(name="Calibri", bold=False, size=9, color="FFFFFF")
        cell.fill  = BLACK_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = none_border

    def section_title(ws, r, last_col, text):
        ws.row_dimensions[r].height = 14
        plain(ws.cell(r, 2), text, size=8, color="888888", border=bot_medium)
        for col in range(3, last_col + 1):
            ws.cell(r, col).border = bot_medium

    ws.column_dimensions["A"].width = 1
    ws.column_dimensions["B"].width = 9
    for col in range(3, T + 4):
        ws.column_dimensions[get_column_letter(col)].width = 5.5
    last_col = T + 2

    # ── Title ──────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 26
    ws.merge_cells(f"B1:{get_column_letter(last_col)}1")
    plain(ws.cell(1, 2), "Assignment 5a - Optimal solution (overtime + permanent expansion)", bold=True, size=13)

    # ── Cost summary ───────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 4
    cost_rows = [
        ("Setup cost",          total_setup,                  False),
        ("Holding cost",        total_holding,                False),
        ("Investment cost X",   invest_X,                     False),
        ("Investment cost Y",   invest_Y,                     False),
        ("Overtime cost X",     total_ot_x,                   False),
        ("Overtime cost Y",     total_ot_y,                   False),
        ("Total overtime cost", total_ot_x + total_ot_y,      False),
        ("Total cost",          model.ObjVal,                 True),
    ]
    for r, (label, val, bold) in enumerate(cost_rows, start=3):
        ws.row_dimensions[r].height = 16
        plain(ws.cell(r, 2), label, size=9, color="555555")
        plain(ws.cell(r, 3), round(clean_num(val), 2), bold=bold, size=9, fmt='"€"#,##0.00')
        ws.merge_cells(f"C{r}:{get_column_letter(last_col)}{r}")

    # ── Expansion & overtime parameters ───────────────────────────────────
    ws.row_dimensions[11].height = 6
    param_rows = [
        ("X  base capacity",     f"{CAP_X_BASE} units / week"),
        ("X  permanent expansion", f"+{round(dx_val)} units  @ €{COST_EXP_X} / unit  (max {CAP_X_MAX_EXP})"),
        ("X  new base capacity",  f"{round(cap_x_new)} units / week"),
        ("X  overtime max",       f"{CAP_X_OT_MAX} units / week  @ €{COST_OT_X} / unit"),
        ("Y  base capacity",      f"{CAP_Y_BASE} min / week"),
        ("Y  permanent expansion", f"+{round(dy_pct_val, 1)}%  @ €{COST_EXP_Y_PCT:,} / %  (max {CAP_Y_MAX_PCT}%)"),
        ("Y  new base capacity",   f"{round(cap_y_new)} min / week"),
        ("Y  overtime max",        f"{CAP_Y_OT_MAX} h / week  @ €{COST_OT_Y} / hour"),
    ]
    for r, (label, val) in enumerate(param_rows, start=12):
        ws.row_dimensions[r].height = 16
        plain(ws.cell(r, 2), label, size=9, color="555555")
        plain(ws.cell(r, 3), val, size=9, bold=("new base" in label))
        ws.merge_cells(f"C{r}:{get_column_letter(last_col)}{r}")

    # ══════════════════════════════════════════════════════════════════════
    # Production schedule
    # ══════════════════════════════════════════════════════════════════════
    r = 21
    section_title(ws, r, last_col, "Production / order schedule (units)")
    r += 1
    ws.row_dimensions[r].height = 16
    hdr(ws.cell(r, 2), "Part")
    for t in periods:
        hdr(ws.cell(r, t+2), str(t))

    for i in parts:
        r += 1
        ws.row_dimensions[r].height = 15
        plain(ws.cell(r, 2), i, size=9, border=bot_thin)
        for t in periods:
            val = round(x[i,t].X) if x[i,t].X > 0.5 else ""
            plain(ws.cell(r, t+2), val, bold=bool(val), size=9, align="center",
                  fmt='#,##0' if val != "" else None, border=bot_thin)

    # ══════════════════════════════════════════════════════════════════════
    # Inventory levels
    # ══════════════════════════════════════════════════════════════════════
    r += 2
    section_title(ws, r, last_col, "Inventory levels (end of period)")
    r += 1
    ws.row_dimensions[r].height = 16
    hdr(ws.cell(r, 2), "Part")
    for t in periods:
        hdr(ws.cell(r, t+2), str(t))

    for i in parts:
        r += 1
        ws.row_dimensions[r].height = 15
        plain(ws.cell(r, 2), i, size=9, border=bot_thin)
        for t in periods:
            v = round(clean_num(I[i,t].X))
            plain(ws.cell(r, t+2), v, size=9, align="center",
                  color="000000" if v > 0 else "BBBBBB",
                  fmt='#,##0', border=bot_thin)

    # ══════════════════════════════════════════════════════════════════════
    # Overtime usage
    # ══════════════════════════════════════════════════════════════════════
    r += 2
    section_title(ws, r, last_col, "Overtime usage per week")
    r += 1
    ws.row_dimensions[r].height = 16
    hdr(ws.cell(r, 2), "")
    for t in periods:
        hdr(ws.cell(r, t+2), str(t))

    for label, fn, fmt in [
        ("X  (units OT)", lambda t: round(clean_num(ot_x[t].X)),          '#,##0'),
        ("Y  (hours OT)", lambda t: round(clean_num(ot_y[t].X), 2),       '0.00'),
        ("OT cost (EUR)", lambda t: round(clean_num(COST_OT_X*ot_x[t].X + COST_OT_Y*ot_y[t].X), 2), '"€"#,##0.00'),
    ]:
        r += 1
        ws.row_dimensions[r].height = 15
        plain(ws.cell(r, 2), label, size=9, border=bot_thin)
        for t in periods:
            v = fn(t)
            col = "CC0000" if v else "BBBBBB"
            plain(ws.cell(r, t+2), v if v else "", bold=bool(v), size=9,
                  align="center", color=col,
                  fmt=fmt if v else None, border=bot_thin)

    # ══════════════════════════════════════════════════════════════════════
    # Capacity utilisation (base + expansion + overtime)
    # ══════════════════════════════════════════════════════════════════════
    r += 2
    section_title(ws, r, last_col,
                  f"Capacity utilisation  (X base: {round(cap_x_new)} u/wk  |  Y base: {round(cap_y_new)} min/wk)")
    r += 1
    ws.row_dimensions[r].height = 16
    hdr(ws.cell(r, 2), "")
    for t in periods:
        hdr(ws.cell(r, t+2), str(t))

    cap_rows = [
        ("X  (units)",   lambda t: sum(PROC_X.get(i,0)*x[i,t].X for i in parts),
                         lambda t: cap_x_new + clean_num(ot_x[t].X), '#,##0'),
        ("Y  (min)",     lambda t: sum(PROC_Y.get(i,0)*x[i,t].X for i in parts),
                         lambda t: cap_y_new + 60*clean_num(ot_y[t].X), '#,##0'),
        ("X  (% used)",  lambda t: sum(PROC_X.get(i,0)*x[i,t].X for i in parts) / (cap_x_new + clean_num(ot_x[t].X)) if (cap_x_new + clean_num(ot_x[t].X)) else 0,
                         None, '0%'),
        ("Y  (% used)",  lambda t: sum(PROC_Y.get(i,0)*x[i,t].X for i in parts) / (cap_y_new + 60*clean_num(ot_y[t].X)) if (cap_y_new + 60*clean_num(ot_y[t].X)) else 0,
                         None, '0%'),
    ]
    for label, load_fn, cap_fn, fmt in cap_rows:
        r += 1
        ws.row_dimensions[r].height = 15
        plain(ws.cell(r, 2), label, size=9, border=bot_thin)
        for t in periods:
            raw = load_fn(t)
            is_pct = "%" in label
            v = round(raw, 3) if is_pct else round(raw)
            cap = (cap_fn(t) if cap_fn else 1)
            pct = raw if is_pct else (raw / cap if cap else 0)
            over = pct > 0.95
            col  = "CC0000" if over else ("000000" if raw > 0 else "BBBBBB")
            plain(ws.cell(r, t+2), v if raw > 0 else "", bold=over, size=9,
                  align="center", color=col,
                  fmt=fmt if raw > 0 else None, border=bot_thin)

    # ══════════════════════════════════════════════════════════════════════
    # Setup decisions
    # ══════════════════════════════════════════════════════════════════════
    r += 2
    section_title(ws, r, last_col, "Setup decisions")
    r += 1
    ws.row_dimensions[r].height = 16
    hdr(ws.cell(r, 2), "Part")
    for t in periods:
        hdr(ws.cell(r, t+2), str(t))

    for i in parts:
        r += 1
        ws.row_dimensions[r].height = 15
        plain(ws.cell(r, 2), i, size=9, border=bot_thin)
        for t in periods:
            plain(ws.cell(r, t+2), "x" if y[i,t].X > 0.5 else "",
                  bold=True, size=9, align="center", border=bot_thin)

    wb.save(OUTPUT_FILE)
    print(f"Results written to {OUTPUT_FILE} -> sheet '{SHEET_NAME}'")
    print(f"Total cost:      EUR {clean_num(model.ObjVal):,.2f}")
    print(f"  Setup:         EUR {clean_num(total_setup):,.2f}")
    print(f"  Holding:       EUR {clean_num(total_holding):,.2f}")
    print(f"  Investment X:  EUR {invest_X:,.2f}  (+{round(dx_val)} units)")
    print(f"  Investment Y:  EUR {invest_Y:,.2f}  (+{round(dy_pct_val,1)}%)")
    print(f"  Overtime X:    EUR {clean_num(total_ot_x):,.2f}")
    print(f"  Overtime Y:    EUR {clean_num(total_ot_y):,.2f}")
