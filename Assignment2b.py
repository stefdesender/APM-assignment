"""
Assignment 2b
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from input_data import (
    PARTS, END_PRODUCT, T, BOM, LEAD_TIME,
    INIT_INV, SETUP_COST, HOLDING_COST,
    DEMAND_REALIZED, BACKORDER_COST
)

# ── Re-solve 2a to get schedule ────────────────────────────────────────────
import gurobipy as gp
from gurobipy import GRB
from input_data import DEMAND_FORECAST, MIN_LOT

def get_parents(part):
    parents = {}
    for parent, children in BOM.items():
        if part in children:
            parents[parent] = children[part]
    return parents

CAPACITY_X = 800
CAPACITY_Y = 7 * 24 * 60 - 80
periods = range(1, T + 1)

model = gp.Model("APM_2a_reload")
model.setParam("OutputFlag", 0)
x = model.addVars(PARTS, periods, name="x", vtype=GRB.INTEGER, lb=0)
y = model.addVars(PARTS, periods, name="y", vtype=GRB.BINARY)
I = model.addVars(PARTS, periods, name="I", lb=0)
BIG_M = {i: sum(DEMAND_FORECAST) * 25 for i in PARTS}

model.setObjective(
    gp.quicksum(SETUP_COST[i]*y[i,t] + HOLDING_COST[i]*I[i,t]
                for i in PARTS for t in periods), GRB.MINIMIZE)

for i in PARTS:
    for t in periods:
        inv_prev = INIT_INV[i] if t == 1 else I[i, t-1]
        op = t - LEAD_TIME[i]
        receipts = x[i, op] if op >= 1 else 0
        int_dem  = gp.quicksum(BOM[p][i]*x[p,t] for p in get_parents(i))
        ext_dem  = DEMAND_FORECAST[t-1] if i == END_PRODUCT else 0
        model.addConstr(I[i,t] == inv_prev + receipts - int_dem - ext_dem)
        model.addConstr(x[i,t] >= MIN_LOT[i] * y[i,t])
        model.addConstr(x[i,t] <= BIG_M[i] * y[i,t])

for t in periods:
    model.addConstr(x['E2801', t] <= CAPACITY_X)
    model.addConstr(3*x['B1401',t] + 2*x['B2302',t] <= CAPACITY_Y)

model.optimize()
schedule = {i: {t: x[i,t].X for t in periods} for i in PARTS}
setup_cost_2a = sum(SETUP_COST[i]*y[i,t].X for i in PARTS for t in periods)

# ── Simulate with realized demand ─────────────────────────────────────────
inventory  = {i: float(INIT_INV[i]) for i in PARTS}
backorders = {t: 0.0 for t in periods}
delivered  = {t: 0.0 for t in periods}
inv_hist   = {i: {} for i in PARTS}   # track inventory per period per part

holding_cost         = 0.0
setup_cost           = 0.0
backorder_cost_total = 0.0
periods_fully_met    = 0

for t in periods:
    for i in PARTS:
        op = t - LEAD_TIME[i]
        qty = schedule[i].get(op, 0.0) if op >= 1 else 0.0
        inventory[i] += qty
        if qty > 0:
            setup_cost += SETUP_COST[i]

    for parent, children in BOM.items():
        prod_qty = schedule[parent].get(t, 0.0)
        if prod_qty > 0:
            for child, qty_per_unit in children.items():
                inventory[child] -= prod_qty * qty_per_unit

    d = DEMAND_REALIZED[t - 1]
    available = inventory[END_PRODUCT]
    if available >= d:
        delivered[t] = d
        inventory[END_PRODUCT] -= d
        periods_fully_met += 1
    else:
        delivered[t] = available
        backorders[t] = d - available
        inventory[END_PRODUCT] = 0

    for i in PARTS:
        if inventory[i] > 0:
            holding_cost += HOLDING_COST[i] * inventory[i]
        inv_hist[i][t] = inventory[i]

    backorder_cost_total += BACKORDER_COST * backorders[t]

total_demand    = sum(DEMAND_REALIZED)
total_delivered = sum(delivered.values())
total_backorder = sum(backorders.values())
service_level   = periods_fully_met / T * 100
fill_rate       = total_delivered / total_demand * 100
total_cost      = setup_cost + holding_cost + backorder_cost_total

# ── Try loading 1b results for comparison ────────────────────────────────
try:
    wb_out = load_workbook("OUTPUT.xlsx", data_only=True)
    ws_1b  = wb_out["Output_1b"]
    cost_1b_total   = ws_1b.cell(6, 3).value   # row 6 = total cost 1b
    cost_1b_setup   = ws_1b.cell(3, 3).value
    cost_1b_holding = ws_1b.cell(4, 3).value
    cost_1b_bo      = ws_1b.cell(5, 3).value
    # service metrics are text strings in 1b sheet; parse from the cell
    sl_1b_str = ws_1b.cell(11, 3).value or ""
    fr_1b_str = ws_1b.cell(12, 3).value or ""
    import re
    sl_1b = float(re.search(r"[\d.]+", str(sl_1b_str)).group()) if sl_1b_str else None
    fr_1b = float(re.search(r"[\d.]+", str(fr_1b_str)).group()) if fr_1b_str else None
    has_1b = True
except Exception:
    has_1b = False

# ── Write to OUTPUT.xlsx ───────────────────────────────────────────────────
OUTPUT_FILE = "OUTPUT.xlsx"
SHEET_NAME  = "Output_2b"

if os.path.exists(OUTPUT_FILE):
    wb = load_workbook(OUTPUT_FILE)
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)
else:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

# ── Styles ────────────────────────────────────────────────────────────────
NO_FILL    = PatternFill(fill_type=None)
BLACK_FILL = PatternFill("solid", start_color="000000", end_color="000000")
none_border = Border()
bot_medium  = Border(bottom=Side(style="medium", color="000000"))
bot_thin    = Border(bottom=Side(style="thin",   color="CCCCCC"))

def plain(cell, val, bold=False, align="left", size=9, color="000000", fmt=None, border=None):
    cell.value = val
    cell.font  = Font(name="Calibri", bold=bold, size=size, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.fill  = NO_FILL
    if border: cell.border = border
    if fmt:    cell.number_format = fmt

def hdr(cell, val):
    cell.value = val
    cell.font  = Font(name="Calibri", bold=False, size=9, color="FFFFFF")
    cell.fill  = BLACK_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = none_border

def section_title(ws, r, last_col, text):
    ws.row_dimensions[r].height = 14
    plain(ws.cell(r, 2), text, size=8, color="888888", border=bot_medium)
    for col in range(3, last_col + 1):
        ws.cell(r, col).border = bot_medium

# ── Column widths ──────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 1
ws.column_dimensions["B"].width = 9
for col in range(3, T + 4):
    ws.column_dimensions[get_column_letter(col)].width = 5.5

last_col = T + 2

# ── Title ──────────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 26
ws.merge_cells(f"B1:{get_column_letter(last_col)}1")
plain(ws.cell(1, 2), "Assignment 2b - Realized demand evaluation (finite capacity)", bold=True, size=13)

# ── Cost summary ───────────────────────────────────────────────────────────
ws.row_dimensions[2].height = 4
for r, (label, val, fmt, bold) in enumerate([
    ("Setup cost",      setup_cost,           '"€"#,##0.00', False),
    ("Holding cost",    holding_cost,          '"€"#,##0.00', False),
    ("Backorder cost",  backorder_cost_total,  '"€"#,##0.00', False),
    ("Total cost",      total_cost,            '"€"#,##0.00', True),
], start=3):
    ws.row_dimensions[r].height = 16
    plain(ws.cell(r, 2), label, size=9, color="555555")
    vc = ws.cell(r, 3)
    plain(vc, round(val, 2), bold=bold, size=9, fmt=fmt)
    ws.merge_cells(f"C{r}:{get_column_letter(last_col)}{r}")

# ── Service metrics ────────────────────────────────────────────────────────
ws.row_dimensions[7].height = 6
ws.row_dimensions[8].height = 14
section_title(ws, 8, last_col, "Service metrics (end product)")

ws.row_dimensions[9].height = 16
plain(ws.cell(9, 2), "Service level", size=9, color="555555")
plain(ws.cell(9, 3),
      f"{service_level:.1f}%  ({periods_fully_met}/{T} periods fully met)", size=9)
ws.merge_cells(f"C9:{get_column_letter(last_col)}9")

ws.row_dimensions[10].height = 16
plain(ws.cell(10, 2), "Fill rate", size=9, color="555555")
plain(ws.cell(10, 3),
      f"{fill_rate:.2f}%  ({total_delivered:,.0f} / {total_demand:,.0f} units delivered)", size=9)
ws.merge_cells(f"C10:{get_column_letter(last_col)}10")

ws.row_dimensions[11].height = 16
plain(ws.cell(11, 2), "Total backordered", size=9, color="555555")
plain(ws.cell(11, 3), f"{total_backorder:,.0f} units", size=9,
      color="CC0000" if total_backorder > 0 else "000000")
ws.merge_cells(f"C11:{get_column_letter(last_col)}11")

# ══════════════════════════════════════════════════════════════════════════
# Comparison table 1b vs 2b
# ══════════════════════════════════════════════════════════════════════════
r = 13
section_title(ws, r, last_col, "Comparison: 1b (infinite capacity) vs 2b (finite capacity)")
r += 1
ws.row_dimensions[r].height = 16

# Use 5 fixed columns for the comparison table
COL_METRIC = 2
COL_1B     = 3
COL_2B     = 4
COL_DIFF   = 5
ws.column_dimensions[get_column_letter(COL_METRIC)].width = 22
ws.column_dimensions[get_column_letter(COL_1B)].width     = 16
ws.column_dimensions[get_column_letter(COL_2B)].width     = 16
ws.column_dimensions[get_column_letter(COL_DIFF)].width   = 16

hdr(ws.cell(r, COL_METRIC), "Metric")
hdr(ws.cell(r, COL_1B),     "1b")
hdr(ws.cell(r, COL_2B),     "2b")
hdr(ws.cell(r, COL_DIFF),   "Difference")

rows_cmp = [
    ("Total cost (EUR)",      cost_1b_total   if has_1b else None, total_cost,            '"€"#,##0.00'),
    ("Setup cost (EUR)",      cost_1b_setup   if has_1b else None, setup_cost,            '"€"#,##0.00'),
    ("Holding cost (EUR)",    cost_1b_holding if has_1b else None, holding_cost,          '"€"#,##0.00'),
    ("Backorder cost (EUR)",  cost_1b_bo      if has_1b else None, backorder_cost_total,  '"€"#,##0.00'),
    ("Service level (%)",     sl_1b           if has_1b else None, service_level,         '0.0'),
    ("Fill rate (%)",         fr_1b           if has_1b else None, fill_rate,             '0.00'),
]

for idx, (label, v1b, v2b, fmt) in enumerate(rows_cmp, start=1):
    r += 1
    ws.row_dimensions[r].height = 15
    plain(ws.cell(r, COL_METRIC), label, size=9, border=bot_thin)

    if v1b is not None:
        plain(ws.cell(r, COL_1B), round(v1b, 2), size=9, align="right", fmt=fmt, border=bot_thin)
        diff_val = round(v2b - v1b, 2)
        diff_col = "CC0000" if diff_val > 0 else ("000000" if diff_val == 0 else "000000")
        diff_fmt = '"+"' + fmt + ';"- "' + fmt if "EUR" not in label else '"+€"#,##0.00;"-€"#,##0.00'
        plain(ws.cell(r, COL_DIFF), diff_val, size=9, align="right",
              color=diff_col, fmt=diff_fmt, border=bot_thin)
    else:
        plain(ws.cell(r, COL_1B), "n/a", size=9, align="right", color="AAAAAA", border=bot_thin)
        plain(ws.cell(r, COL_DIFF), "", size=9, border=bot_thin)

    plain(ws.cell(r, COL_2B), round(v2b, 2), size=9, align="right", fmt=fmt, border=bot_thin)

# ══════════════════════════════════════════════════════════════════════════
# Production schedule
# ══════════════════════════════════════════════════════════════════════════
r += 2
section_title(ws, r, last_col, "Production / order schedule (units)")
r += 1
ws.row_dimensions[r].height = 16
hdr(ws.cell(r, 2), "Part")
for t in periods:
    hdr(ws.cell(r, t + 2), str(t))

for i in PARTS:
    r += 1
    ws.row_dimensions[r].height = 15
    plain(ws.cell(r, 2), i, size=9, border=bot_thin)
    for t in periods:
        val = round(schedule[i][t]) if schedule[i][t] > 0.5 else ""
        plain(ws.cell(r, t + 2), val, bold=bool(val), size=9, align="center",
              fmt='#,##0' if val != "" else None, border=bot_thin)

# ══════════════════════════════════════════════════════════════════════════
# Inventory levels (realized)
# ══════════════════════════════════════════════════════════════════════════
r += 2
section_title(ws, r, last_col, "Inventory levels (end of period, realized)")
r += 1
ws.row_dimensions[r].height = 16
hdr(ws.cell(r, 2), "Part")
for t in periods:
    hdr(ws.cell(r, t + 2), str(t))

for i in PARTS:
    r += 1
    ws.row_dimensions[r].height = 15
    plain(ws.cell(r, 2), i, size=9, border=bot_thin)
    for t in periods:
        v = round(inv_hist[i][t])
        col = "CC0000" if v < 0 else ("BBBBBB" if v == 0 else "000000")
        plain(ws.cell(r, t + 2), v, size=9, align="center",
              color=col, fmt='#,##0;-#,##0', border=bot_thin)

# ══════════════════════════════════════════════════════════════════════════
# Backorder schedule (end product only)
# ══════════════════════════════════════════════════════════════════════════
r += 2
section_title(ws, r, last_col, f"Backorder schedule - {END_PRODUCT} (end product only)")
r += 1
ws.row_dimensions[r].height = 16
hdr(ws.cell(r, 2), "Part")
for t in periods:
    hdr(ws.cell(r, t + 2), str(t))

r += 1
ws.row_dimensions[r].height = 15
plain(ws.cell(r, 2), END_PRODUCT, size=9, border=bot_thin)
for t in periods:
    v = round(backorders[t])
    if v > 0:
        plain(ws.cell(r, t + 2), v, bold=True, size=9, align="center",
              color="CC0000", fmt='#,##0', border=bot_thin)
    else:
        plain(ws.cell(r, t + 2), "", size=9, align="center", border=bot_thin)

# ══════════════════════════════════════════════════════════════════════════
# Demand vs delivered
# ══════════════════════════════════════════════════════════════════════════
r += 2
section_title(ws, r, last_col, "Demand vs delivered (end product)")
r += 1
ws.row_dimensions[r].height = 16
hdr(ws.cell(r, 2), "")
for t in periods:
    hdr(ws.cell(r, t + 2), str(t))

r += 1
ws.row_dimensions[r].height = 15
plain(ws.cell(r, 2), "Demand", size=9, border=bot_thin)
for t in periods:
    plain(ws.cell(r, t + 2), DEMAND_REALIZED[t-1], size=9, align="center",
          fmt='#,##0', border=bot_thin)

r += 1
ws.row_dimensions[r].height = 15
plain(ws.cell(r, 2), "Delivered", size=9, border=bot_thin)
for t in periods:
    v = round(delivered[t])
    shortage = v < DEMAND_REALIZED[t-1]
    plain(ws.cell(r, t + 2), v, size=9, align="center",
          color="CC0000" if shortage else "000000",
          fmt='#,##0', border=bot_thin)

# ══════════════════════════════════════════════════════════════════════════
# Setup decisions
# ══════════════════════════════════════════════════════════════════════════
r += 2
section_title(ws, r, last_col, "Setup decisions")
r += 1
ws.row_dimensions[r].height = 16
hdr(ws.cell(r, 2), "Part")
for t in periods:
    hdr(ws.cell(r, t + 2), str(t))

for i in PARTS:
    r += 1
    ws.row_dimensions[r].height = 15
    plain(ws.cell(r, 2), i, size=9, border=bot_thin)
    for t in periods:
        setup = 1 if schedule[i][t] > 0.5 else 0
        plain(ws.cell(r, t + 2), "x" if setup else "",
              bold=True, size=9, align="center", border=bot_thin)

wb.save(OUTPUT_FILE)
print(f"Results written to {OUTPUT_FILE} -> sheet '{SHEET_NAME}'")
print(f"Total cost:      EUR {total_cost:,.2f}")
print(f"  Setup:         EUR {setup_cost:,.2f}")
print(f"  Holding:       EUR {holding_cost:,.2f}")
print(f"  Backorder:     EUR {backorder_cost_total:,.2f}")
print(f"Service level:   {service_level:.1f}%")
print(f"Fill rate:       {fill_rate:.2f}%")