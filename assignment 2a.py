import gurobipy as gp
from gurobipy import GRB
from openpyxl import load_workbook

# ═══════════════════════════════════════════════════════════════════════════
# 1.  READ INPUT DATA FROM EXCEL
# ═══════════════════════════════════════════════════════════════════════════

INPUT_FILE = r"C:\Users\sarev\OneDrive\Documenten\1e master\2e sem\APM\input_data.xlsx"

def load_inputs(path):
    wb = load_workbook(path, data_only=True)

    ws = wb["Parameters"]
    parts_order = []
    params = {}
    for row in ws.iter_rows(min_row=3, max_row=9, min_col=1, max_col=8, values_only=True):
        part, lt, min_lot, inv0, setup, hold, ws_id, proc = row
        parts_order.append(part)
        params[part] = {
            "lt":      int(lt),
            "min_lot": int(min_lot),
            "inv0":    int(inv0),
            "setup":   float(setup),
            "hold":    float(hold),
            "ws":      str(ws_id).strip() if ws_id else None,
            "proc":    float(proc) if proc not in ("", None) else 0.0,
        }

    bom = {}
    for row in ws.iter_rows(min_row=13, max_row=18, min_col=1, max_col=3, values_only=True):
        parent, component, qty = row
        bom.setdefault(parent, {})[component] = int(qty)

    cap = {}
    for row in ws.iter_rows(min_row=22, max_row=23, min_col=1, max_col=2, values_only=True):
        ws_name, value = row
        cap[str(ws_name).strip()] = float(value)

    ws2 = wb["Demand"]
    demand_forecast = {}
    demand_realized = {}
    for row in ws2.iter_rows(min_row=3, max_row=32, min_col=1, max_col=3, values_only=True):
        week, fc, re = row
        demand_forecast[int(week)] = int(fc)
        demand_realized[int(week)] = int(re)

    ws3 = wb["Backorder"]
    backorder_cost = float(ws3.cell(3, 2).value)

    wb.close()
    return parts_order, params, bom, cap, demand_forecast, demand_realized, backorder_cost


# ═══════════════════════════════════════════════════════════════════════════
# 2.  MIP MODEL
# ═══════════════════════════════════════════════════════════════════════════

def build_and_solve(parts, params, bom, cap, demand_dict, T,
                    label="forecast", backorder_cost=None):
    periods = list(range(1, T + 1))
    BIG_M   = 1_000_000

    model = gp.Model(f"APM_Assign2_{label}")
    model.Params.LogToConsole = 1

    x, y, I, B = {}, {}, {}, {}
    for p in parts:
        for t in periods:
            x[p, t] = model.addVar(lb=0, vtype=GRB.INTEGER,    name=f"x_{p}_{t}")
            y[p, t] = model.addVar(vtype=GRB.BINARY,           name=f"y_{p}_{t}")
            I[p, t] = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name=f"I_{p}_{t}")
            if backorder_cost is not None:
                B[p, t] = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name=f"B_{p}_{t}")

    obj = (gp.quicksum(params[p]["setup"] * y[p, t] for p in parts for t in periods)
         + gp.quicksum(params[p]["hold"]  * I[p, t] for p in parts for t in periods))
    if backorder_cost is not None:
        obj += gp.quicksum(backorder_cost * B["E2801", t] for t in periods)
    model.setObjective(obj, GRB.MINIMIZE)

    for p in parts:
        lt = params[p]["lt"]
        for t in periods:
            inv_prev  = params[p]["inv0"] if t == 1 else I[p, t - 1]
            prod_recv = x[p, t - lt] if (t - lt) >= 1 else 0

            if p == "E2801":
                use = demand_dict[t]
                if backorder_cost is not None:
                    back_prev = B[p, t - 1] if t > 1 else 0
                    model.addConstr(
                        inv_prev + prod_recv + back_prev - use == I[p, t] - B[p, t],
                        name=f"inv_bal_{p}_{t}")
                else:
                    model.addConstr(inv_prev + prod_recv - use == I[p, t],
                                    name=f"inv_bal_{p}_{t}")
            else:
                use_expr = gp.quicksum(
                    children[p] * x[parent, t]
                    for parent, children in bom.items() if p in children)
                model.addConstr(inv_prev + prod_recv - use_expr == I[p, t],
                                name=f"inv_bal_{p}_{t}")

            model.addConstr(x[p, t] >= params[p]["min_lot"] * y[p, t],
                            name=f"min_lot_{p}_{t}")
            model.addConstr(x[p, t] <= BIG_M * y[p, t],
                            name=f"setup_link_{p}_{t}")

    cap_X = cap.get("X")
    cap_Y = cap.get("Y_weekly_minutes")
    ws_X  = [p for p in parts if params[p]["ws"] == "X"]
    ws_Y  = [p for p in parts if params[p]["ws"] == "Y"]

    for t in periods:
        if cap_X and ws_X:
            model.addConstr(gp.quicksum(x[p, t] for p in ws_X) <= cap_X,
                            name=f"cap_X_{t}")
        if cap_Y and ws_Y:
            model.addConstr(
                gp.quicksum(params[p]["proc"] * x[p, t] for p in ws_Y) <= cap_Y,
                name=f"cap_Y_{t}")

    model.optimize()

    if model.Status not in (GRB.OPTIMAL, GRB.SUBOPTIMAL):
        print(f"[{label}] No optimal solution. Status = {model.Status}")
        return model, None

    sol = {
        "status": model.Status, "obj": model.ObjVal,
        "periods": periods, "parts": parts,
        "x": {}, "y": {}, "I": {}, "B": {},
        "cap_X": cap_X, "cap_Y": cap_Y, "demand": demand_dict,
        "setup_cost": 0.0, "holding_cost": 0.0, "backorder_cost_total": 0.0,
    }
    for p in parts:
        sol["x"][p] = {}; sol["y"][p] = {}; sol["I"][p] = {}; sol["B"][p] = {}
        for t in periods:
            sol["x"][p][t] = round(x[p, t].X)
            sol["y"][p][t] = int(round(y[p, t].X))
            sol["I"][p][t] = round(I[p, t].X, 2)
            bval = round(B[p, t].X, 2) if (backorder_cost and p == "E2801") else 0.0
            sol["B"][p][t] = bval
            sol["setup_cost"]   += params[p]["setup"] * sol["y"][p][t]
            sol["holding_cost"] += params[p]["hold"]  * sol["I"][p][t]

    if backorder_cost:
        sol["backorder_cost_total"] = sum(
            backorder_cost * sol["B"]["E2801"][t] for t in periods)
        no_back = sum(1 for t in periods if sol["B"]["E2801"][t] < 0.5)
        sol["service_level"] = no_back / len(periods) * 100
        tot_dem = sum(demand_dict[t] for t in periods)
        tot_bk  = sum(sol["B"]["E2801"][t] for t in periods)
        sol["fill_rate"] = (1 - tot_bk / tot_dem) * 100

    sol["util_X"] = {t: sum(sol["x"][p][t] for p in ws_X) for t in periods}
    sol["util_Y"] = {t: sum(params[p]["proc"] * sol["x"][p][t]
                            for p in ws_Y) for t in periods}
    return model, sol


# ═══════════════════════════════════════════════════════════════════════════
# 3.  PRINT RESULTS TO CONSOLE
# ═══════════════════════════════════════════════════════════════════════════

def print_results(sol_a, sol_b):
    SEP = "=" * 80

    def print_plan(sol, title):
        print(f"\n{SEP}")
        print(f"  {title}")
        print(SEP)
        parts   = sol["parts"]
        periods = sol["periods"]
        has_bk  = any(sol["B"].get("E2801", {}).get(t, 0) > 0 for t in periods)

        # Header
        header = f"{'Period':>6}  {'Demand':>7}"
        for p in parts:
            header += f"  {p+' Prod':>12}  {p+' Setup':>11}  {p+' Inv':>10}"
            if p == "E2801" and has_bk:
                header += f"  {p+' Backlog':>13}"
        print(header)
        print("-" * len(header))

        for t in periods:
            line = f"{t:>6}  {sol['demand'][t]:>7}"
            for p in parts:
                line += f"  {sol['x'][p][t]:>12}  {sol['y'][p][t]:>11}  {sol['I'][p][t]:>10.2f}"
                if p == "E2801" and has_bk:
                    line += f"  {sol['B']['E2801'].get(t, 0):>13.2f}"
            print(line)

    def print_capacity(sol_a, sol_b):
        print(f"\n{SEP}")
        print("  CAPACITY UTILISATION")
        print(SEP)
        cap_X = int(sol_a["cap_X"])
        cap_Y = int(sol_a["cap_Y"])
        print(f"{'Period':>6}  {'2a WS-X':>10}  {'2a WS-Y':>12}  {'2b WS-X':>10}  {'2b WS-Y':>12}")
        print(f"{'':>6}  {'(cap '+str(cap_X)+')':>10}  {'(cap '+str(cap_Y)+' min)':>12}  {'(cap '+str(cap_X)+')':>10}  {'(cap '+str(cap_Y)+' min)':>12}")
        print("-" * 60)
        for t in sol_a["periods"]:
            ax = sol_a["util_X"][t]; ay = sol_a["util_Y"][t]
            bx = sol_b["util_X"][t]; by = sol_b["util_Y"][t]
            flag = lambda v, cap: " !" if v > cap else "  "
            print(f"{t:>6}  {ax:>9}{flag(ax,cap_X)}  {ay:>11.1f}{flag(ay,cap_Y)}  "
                  f"{bx:>9}{flag(bx,cap_X)}  {by:>11.1f}{flag(by,cap_Y)}")
        print(f"{'Cap':>6}  {cap_X:>10}  {cap_Y:>12}  {cap_X:>10}  {cap_Y:>12}")

    print_plan(sol_a, "2a – FORECAST PLAN")
    print_plan(sol_b, "2b – REALIZED PLAN (with backorders)")
    print_capacity(sol_a, sol_b)

    print(f"\n{SEP}")
    print("  COST SUMMARY")
    print(SEP)
    print(f"  {'':30} {'2a Forecast':>15}  {'2b Realized':>15}")
    print(f"  {'Setup cost (€)':30} {sol_a['setup_cost']:>15,.2f}  {sol_b['setup_cost']:>15,.2f}")
    print(f"  {'Holding cost (€)':30} {sol_a['holding_cost']:>15,.2f}  {sol_b['holding_cost']:>15,.2f}")
    print(f"  {'Backorder cost (€)':30} {'–':>15}  {sol_b['backorder_cost_total']:>15,.2f}")
    print(f"  {'Total objective (€)':30} {sol_a['obj']:>15,.2f}  {sol_b['obj']:>15,.2f}")
    if "service_level" in sol_b:
        print(f"\n  Service level : {sol_b['service_level']:.1f}%")
        print(f"  Fill rate     : {sol_b['fill_rate']:.1f}%")
    print(SEP)


# ═══════════════════════════════════════════════════════════════════════════
# 4.  MAIN
# ═══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("Reading input data from:", INPUT_FILE)
    parts, params, bom, cap, demand_fc, demand_re, back_cost = load_inputs(INPUT_FILE)
    T = 30

    print(f"Capacity X = {cap.get('X')} units/week")
    print(f"Capacity Y = {cap.get('Y_weekly_minutes')} min/week")

    print("\n" + "="*60)
    print("ASSIGNMENT 2a – Finite Capacity, Forecasted Demand")
    print("="*60)
    _, sol_a = build_and_solve(parts, params, bom, cap, demand_fc, T,
                               label="forecast")

    print("\n" + "="*60)
    print("ASSIGNMENT 2b – Finite Capacity, Realized Demand (backordering)")
    print("="*60)
    _, sol_b = build_and_solve(parts, params, bom, cap, demand_re, T,
                               label="realized", backorder_cost=back_cost)

    if sol_a and sol_b:
        print_results(sol_a, sol_b)