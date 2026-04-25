"""
Assignment 2a
"""

import gurobipy as gp
from gurobipy import GRB
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from input_data import (
    PARTS, END_PRODUCT, T, BOM, LEAD_TIME, MIN_LOT,
    INIT_INV, SETUP_COST, HOLDING_COST, DEMAND_FORECAST
)

def get_parents(part):
    parents = {}
    for parent, children in BOM.items():
        if part in children:
            parents[parent] = children[part]
    return parents

CAPACITY_X = 800
CAPACITY_Y = 7 * 24 * 60 - 80

model = gp.Model("APM_Assignment2a")
periods = range(1, T + 1)
parts   = PARTS

x = model.addVars(parts, periods, name="x", vtype=GRB.INTEGER, lb=0)
y = model.addVars(parts, periods, name="y", vtype=GRB.BINARY)
I = model.addVars(parts, periods, name="I", lb=0)

BIG_M = {i: sum(DEMAND_FORECAST) * 25 for i in parts}

model.setObjective(
    gp.quicksum(SETUP_COST[i] * y[i,t] + HOLDING_COST[i] * I[i,t]
                for i in parts for t in periods), GRB.MINIMIZE)

for i in parts:
    for t in periods:
        inv_prev = INIT_INV[i] if t == 1 else I[i, t-1]
        op = t - LEAD_TIME[i]
        receipts = x[i, op] if op >= 1 else 0
        int_dem  = gp.quicksum(BOM[p][i] * x[p,t] for p in get_parents(i))
        ext_dem  = DEMAND_FORECAST[t-1] if i == END_PRODUCT else 0
        model.addConstr(I[i,t] == inv_prev + receipts - int_dem - ext_dem,
                        name=f"inv_balance_{i}_{t}")
        model.addConstr(x[i,t] >= MIN_LOT[i] * y[i,t], name=f"min_lot_{i}_{t}")
        model.addConstr(x[i,t] <= BIG_M[i] * y[i,t],   name=f"bigM_{i}_{t}")

for t in periods:
    model.addConstr(x['E2801', t] <= CAPACITY_X, name=f"capacity_X_{t}")
    model.addConstr(3 * x['B1401', t] + 2 * x['B2302', t] <= CAPACITY_Y,
                    name=f"capacity_Y_{t}")

model.optimize()

if model.status == GRB.OPTIMAL:
    total_setup   = sum(SETUP_COST[i]   * y[i,t].X for i in parts for t in periods)
    total_holding = sum(HOLDING_COST[i] * I[i,t].X for i in parts for t in periods)

    OUTPUT_FILE = "OUTPUT.xlsx"
    SHEET_NAME  = "Output_2a"

    if os.path.exists(OUTPUT_FILE):
        wb = load_workbook(OUTPUT_FILE)
        if SHEET_NAME in wb.sheetnames:
            del wb[SHEET_NAME]
        ws = wb.create_sheet(SHEET_NAME)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME

    # ── Styles ────────────────────────────────────────────────────────────
    NO_FILL    = PatternFill(fill_type=None)
    BLACK_FILL = PatternFill("solid", start_color="000000", end_color="000000")
    none_border = Border()
    bot_medium  = Border(bottom=Side(style="medium", color="000000"))
    bot_thin    = Border(bottom=Side(style="thin",   color="CCCCCC"))

    def plain(cell, val, bold=False, align="left", size=9, color="000000", fmt=None, border=None):
        cell.value = val
        cell.font  = Font(name="Calibri", bold=bold, size=size, color=color)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.fill  = NO_FILL
        if border: cell.border = border
        if fmt:    cell.number_format = fmt

    def hdr(cell, val):
        cell.value = val
        cell.font  = Font(name="Calibri", bold=False, size=9, color="FFFFFF")
        cell.fill  = BLACK_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = none_border

    def section_title(ws, r, last_col, text):
        ws.row_dimensions[r].height = 14
        plain(ws.cell(r, 2), text, size=8, color="888888", border=bot_medium)
        for col in range(3, last_col + 1):
            ws.cell(r, col).border = bot_medium

    # ── Column widths ──────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 1
    ws.column_dimensions["B"].width = 9
    for col in range(3, T + 4):
        ws.column_dimensions[get_column_letter(col)].width = 5.5

    last_col = T + 2

    # ── Title ──────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 26
    ws.merge_cells(f"B1:{get_column_letter(last_col)}1")
    plain(ws.cell(1, 2), "Assignment 2a - Optimal solution (finite capacity)", bold=True, size=13)

    # ── Cost summary ───────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 4
    for r, (label, val, fmt, bold) in enumerate([
        ("Total cost",   model.ObjVal,  '"€"#,##0.00', True),
        ("Setup cost",   total_setup,   '"€"#,##0.00', False),
        ("Holding cost", total_holding, '"€"#,##0.00', False),
    ], start=3):
        ws.row_dimensions[r].height = 16
        plain(ws.cell(r, 2), label, size=9, color="555555")
        vc = ws.cell(r, 3)
        plain(vc, round(val, 2), bold=bold, size=9, fmt=fmt)
        ws.merge_cells(f"C{r}:{get_column_letter(last_col)}{r}")

    # ── Capacity parameters ────────────────────────────────────────────────
    ws.row_dimensions[6].height = 4
    ws.row_dimensions[7].height = 16
    plain(ws.cell(7, 2), "Workstation X capacity", size=9, color="555555")
    vc = ws.cell(7, 3)
    plain(vc, f"{CAPACITY_X} units / week", size=9)
    ws.merge_cells(f"C7:{get_column_letter(last_col)}7")

    ws.row_dimensions[8].height = 16
    plain(ws.cell(8, 2), "Workstation Y capacity", size=9, color="555555")
    vc = ws.cell(8, 3)
    plain(vc, f"{CAPACITY_Y} min / week  (3 min/B1401, 2 min/B2302)", size=9)
    ws.merge_cells(f"C8:{get_column_letter(last_col)}8")

    # ══════════════════════════════════════════════════════════════════════
    # Production schedule
    # ══════════════════════════════════════════════════════════════════════
    r = 10
    section_title(ws, r, last_col, "Production / order schedule (units)")
    r += 1
    ws.row_dimensions[r].height = 16
    hdr(ws.cell(r, 2), "Part")
    for t in periods:
        hdr(ws.cell(r, t + 2), str(t))

    for idx, i in enumerate(parts, start=1):
        r += 1
        ws.row_dimensions[r].height = 15
        plain(ws.cell(r, 2), i, size=9, border=bot_thin)
        for t in periods:
            val = round(x[i,t].X) if x[i,t].X > 0.5 else ""
            plain(ws.cell(r, t + 2), val, bold=bool(val), size=9, align="center",
                  fmt='#,##0' if val != "" else None, border=bot_thin)

    # ══════════════════════════════════════════════════════════════════════
    # Inventory levels
    # ══════════════════════════════════════════════════════════════════════
    r += 2
    section_title(ws, r, last_col, "Inventory levels (end of period)")
    r += 1
    ws.row_dimensions[r].height = 16
    hdr(ws.cell(r, 2), "Part")
    for t in periods:
        hdr(ws.cell(r, t + 2), str(t))

    for idx, i in enumerate(parts, start=1):
        r += 1
        ws.row_dimensions[r].height = 15
        plain(ws.cell(r, 2), i, size=9, border=bot_thin)
        for t in periods:
            v = round(I[i,t].X)
            plain(ws.cell(r, t + 2), v, size=9, align="center",
                  color="000000" if v > 0 else "BBBBBB",
                  fmt='#,##0', border=bot_thin)

    # ══════════════════════════════════════════════════════════════════════
    # Capacity usage
    # ══════════════════════════════════════════════════════════════════════
    r += 2
    section_title(ws, r, last_col, "Capacity usage per week")
    r += 1
    ws.row_dimensions[r].height = 16
    hdr(ws.cell(r, 2), "")
    for t in periods:
        hdr(ws.cell(r, t + 2), str(t))

    # Workstation X
    r += 1
    ws.row_dimensions[r].height = 15
    plain(ws.cell(r, 2), "X  (units)", size=9, border=bot_thin)
    for t in periods:
        used = round(x['E2801', t].X)
        pct  = used / CAPACITY_X
        col  = "CC0000" if pct > 0.95 else ("000000" if used > 0 else "BBBBBB")
        plain(ws.cell(r, t + 2), used if used else "", bold=(pct > 0.95),
              size=9, align="center", color=col,
              fmt='#,##0' if used else None, border=bot_thin)

    # Workstation Y
    r += 1
    ws.row_dimensions[r].height = 15
    plain(ws.cell(r, 2), "Y  (min)", size=9, border=bot_thin)
    for t in periods:
        used = round(3 * x['B1401', t].X + 2 * x['B2302', t].X)
        pct  = used / CAPACITY_Y
        col  = "CC0000" if pct > 0.95 else ("000000" if used > 0 else "BBBBBB")
        plain(ws.cell(r, t + 2), used if used else "", bold=(pct > 0.95),
              size=9, align="center", color=col,
              fmt='#,##0' if used else None, border=bot_thin)

    # Capacity %
    r += 1
    ws.row_dimensions[r].height = 15
    plain(ws.cell(r, 2), "X  (% used)", size=9, border=bot_thin)
    for t in periods:
        pct = x['E2801', t].X / CAPACITY_X
        col = "CC0000" if pct > 0.95 else ("000000" if pct > 0 else "BBBBBB")
        plain(ws.cell(r, t + 2), round(pct, 3) if pct > 0 else "",
              size=9, align="center", color=col,
              fmt='0%' if pct > 0 else None, border=bot_thin)

    r += 1
    ws.row_dimensions[r].height = 15
    plain(ws.cell(r, 2), "Y  (% used)", size=9, border=bot_thin)
    for t in periods:
        used = 3 * x['B1401', t].X + 2 * x['B2302', t].X
        pct  = used / CAPACITY_Y
        col  = "CC0000" if pct > 0.95 else ("000000" if pct > 0 else "BBBBBB")
        plain(ws.cell(r, t + 2), round(pct, 3) if pct > 0 else "",
              size=9, align="center", color=col,
              fmt='0%' if pct > 0 else None, border=bot_thin)

    # ══════════════════════════════════════════════════════════════════════
    # Setup decisions
    # ══════════════════════════════════════════════════════════════════════
    r += 2
    section_title(ws, r, last_col, "Setup decisions")
    r += 1
    ws.row_dimensions[r].height = 16
    hdr(ws.cell(r, 2), "Part")
    for t in periods:
        hdr(ws.cell(r, t + 2), str(t))

    for idx, i in enumerate(parts, start=1):
        r += 1
        ws.row_dimensions[r].height = 15
        plain(ws.cell(r, 2), i, size=9, border=bot_thin)
        for t in periods:
            setup = int(round(y[i,t].X))
            plain(ws.cell(r, t + 2), "x" if setup else "",
                  bold=True, size=9, align="center", border=bot_thin)

    wb.save(OUTPUT_FILE)
    print(f"Results written to {OUTPUT_FILE} -> sheet '{SHEET_NAME}'")
    print(f"Total cost:    EUR {model.ObjVal:,.2f}")
    print(f"  Setup:       EUR {total_setup:,.2f}")
    print(f"  Holding:     EUR {total_holding:,.2f}")
    