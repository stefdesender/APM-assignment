"""
APM Project 2026 - Assignment 4b
"""

import gurobipy as gp
from gurobipy import GRB
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from input_data import (
    PARTS, END_PRODUCT, T, BOM, LEAD_TIME, MIN_LOT,
    INIT_INV, SETUP_COST, HOLDING_COST,
    DEMAND_FORECAST, DEMAND_REALIZED, BACKORDER_COST
)

CAP_X_BASE     = 800
PROC_X         = {END_PRODUCT: 1}
CAP_Y_BASE     = 60 * 24 * 7 - 80
PROC_Y         = {"B1401": 3, "B2302": 2}
COST_EXP_X     = 10
COST_EXP_Y_PCT = 1_500

def get_parents(part):
    parents = {}
    for parent, children in BOM.items():
        if part in children:
            parents[parent] = children[part]
    return parents

def clean(val):
    return 0 if abs(val) < 1e-6 else round(val)

# ── Re-solve 4a to get fixed plan ─────────────────────────────────────────
from input_data import DEMAND_FORECAST as DF
from input_data import MIN_LOT

CAP_X_MAX_EXP  = 200
CAP_Y_MAX_PCT  = 40
periods = range(1, T + 1)

m4a = gp.Model("APM_4a_reload")
m4a.setParam("OutputFlag", 0)
xa     = m4a.addVars(PARTS, periods, name="x",  vtype=GRB.INTEGER, lb=0)
ya     = m4a.addVars(PARTS, periods, name="y",  vtype=GRB.BINARY)
Ia     = m4a.addVars(PARTS, periods, name="I",  vtype=GRB.INTEGER, lb=0)
dxa    = m4a.addVar(name="dx",     vtype=GRB.INTEGER, lb=0, ub=CAP_X_MAX_EXP)
dya    = m4a.addVar(name="dy_pct", lb=0, ub=CAP_Y_MAX_PCT)
BIG_M  = {i: sum(DF)*25 for i in PARTS}

m4a.setObjective(
    gp.quicksum(SETUP_COST[i]*ya[i,t] + HOLDING_COST[i]*Ia[i,t]
                for i in PARTS for t in periods)
    + COST_EXP_X*dxa + COST_EXP_Y_PCT*dya, GRB.MINIMIZE)

for i in PARTS:
    for t in periods:
        inv_prev = INIT_INV[i] if t==1 else Ia[i,t-1]
        op = t - LEAD_TIME[i]
        rec = xa[i,op] if op>=1 else 0
        int_d = gp.quicksum(BOM[p][i]*xa[p,t] for p in get_parents(i))
        ext_d = DF[t-1] if i==END_PRODUCT else 0
        m4a.addConstr(Ia[i,t] == inv_prev + rec - int_d - ext_d)
        m4a.addConstr(xa[i,t] >= MIN_LOT[i]*ya[i,t])
        m4a.addConstr(xa[i,t] <= BIG_M[i]*ya[i,t])

for t in periods:
    m4a.addConstr(gp.quicksum(PROC_X.get(i,0)*xa[i,t] for i in PARTS) <= CAP_X_BASE + dxa)
    m4a.addConstr(gp.quicksum(PROC_Y.get(i,0)*xa[i,t] for i in PARTS)
                  <= CAP_Y_BASE + (CAP_Y_BASE/100)*dya)
m4a.optimize()

dx_val     = dxa.X
dy_pct_val = dya.X
invest_X   = COST_EXP_X * dx_val
invest_Y   = COST_EXP_Y_PCT * dy_pct_val
cap_x_new  = CAP_X_BASE + dx_val
cap_y_new  = CAP_Y_BASE * (1 + dy_pct_val/100)
cost_4a    = m4a.ObjVal

fixed_x = {i: {t: xa[i,t].X for t in periods} for i in PARTS}
setup_cost_fixed = sum(SETUP_COST[i] for i in PARTS for t in periods if xa[i,t].X > 0.5)

# ── Evaluation model ───────────────────────────────────────────────────────
model = gp.Model("APM_4b")
model.setParam("OutputFlag", 0)

I  = model.addVars(PARTS, periods, name="I",  vtype=GRB.INTEGER, lb=0)
BO = model.addVars(periods,        name="BO", lb=0)

model.setObjective(
    gp.quicksum(HOLDING_COST[i]*I[i,t] for i in PARTS for t in periods)
    + BACKORDER_COST * gp.quicksum(BO[t] for t in periods),
    GRB.MINIMIZE)

for i in PARTS:
    for t in periods:
        inv_prev = INIT_INV[i] if t==1 else I[i,t-1]
        op = t - LEAD_TIME[i]
        rec = fixed_x[i].get(op, 0.0) if op>=1 else 0.0
        int_d = sum(BOM[p][i]*fixed_x[p].get(t,0.0) for p in get_parents(i))
        if i == END_PRODUCT:
            bo_prev = BO[t-1] if t>1 else 0
            model.addConstr(
                I[i,t] - BO[t] == inv_prev - bo_prev + rec - int_d - DEMAND_REALIZED[t-1],
                name=f"inv_{i}_{t}")
        else:
            model.addConstr(I[i,t] == inv_prev + rec - int_d, name=f"inv_{i}_{t}")

model.optimize()

total_holding   = sum(HOLDING_COST[i]*I[i,t].X for i in PARTS for t in periods)
total_backorder = BACKORDER_COST * sum(BO[t].X for t in periods)
total_cost_4b   = setup_cost_fixed + total_holding + total_backorder + invest_X + invest_Y

periods_no_bo = sum(1 for t in periods if BO[t].X < 0.5)
service_level = periods_no_bo / T
total_realized = sum(DEMAND_REALIZED)
new_bo = []
for t in periods:
    bo_tm1 = BO[t-1].X if t>1 else 0.0
    new_bo.append(max(0.0, BO[t].X - bo_tm1))
total_new_bo = sum(new_bo)
fill_rate     = 1.0 - total_new_bo / total_realized

# ── Read reference results from OUTPUT.xlsx ───────────────────────────────
def read_float(ws, row, col=3):
    try:
        return float(ws.cell(row, col).value)
    except Exception:
        return None

import re
def parse_pct(s):
    m = re.search(r"[\d.]+", str(s) if s else "")
    return float(m.group()) if m else None

refs = {}
try:
    wb_ref = load_workbook("OUTPUT.xlsx", data_only=True)
    # 1b: setup r3, holding r4, bo r5, total r6, sl r11, fr r12
    ws1 = wb_ref["Output_1b"]
    refs["1b"] = dict(setup=read_float(ws1,3), holding=read_float(ws1,4),
                      bo=read_float(ws1,5),    total=read_float(ws1,6),
                      sl=parse_pct(ws1.cell(11,3).value),
                      fr=parse_pct(ws1.cell(12,3).value))
    # 2b: setup r3, holding r4, bo r5, total r6, sl r9, fr r10
    ws2 = wb_ref["Output_2b"]
    refs["2b"] = dict(setup=read_float(ws2,3), holding=read_float(ws2,4),
                      bo=read_float(ws2,5),    total=read_float(ws2,6),
                      sl=parse_pct(ws2.cell(9,3).value),
                      fr=parse_pct(ws2.cell(10,3).value))
    # 3b: setup r3, holding r4, ot_x r5, ot_y r6, bo r7, total r8, sl r13, fr r14
    ws3 = wb_ref["Output_3b"]
    refs["3b"] = dict(setup=read_float(ws3,3), holding=read_float(ws3,4),
                      bo=read_float(ws3,7),    total=read_float(ws3,8),
                      sl=parse_pct(ws3.cell(13,3).value),
                      fr=parse_pct(ws3.cell(14,3).value))
    has_ref = True
except Exception:
    has_ref = False

# ── Write to OUTPUT.xlsx ───────────────────────────────────────────────────
OUTPUT_FILE = "OUTPUT.xlsx"
SHEET_NAME  = "Output_4b"

if os.path.exists(OUTPUT_FILE):
    wb = load_workbook(OUTPUT_FILE)
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)
else:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

# ── Styles ────────────────────────────────────────────────────────────────
NO_FILL    = PatternFill(fill_type=None)
BLACK_FILL = PatternFill("solid", start_color="000000", end_color="000000")
none_border = Border()
bot_medium  = Border(bottom=Side(style="medium", color="000000"))
bot_thin    = Border(bottom=Side(style="thin",   color="CCCCCC"))

def plain(cell, val, bold=False, align="left", size=9, color="000000", fmt=None, border=None):
    cell.value = val
    cell.font  = Font(name="Calibri", bold=bold, size=size, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.fill  = NO_FILL
    if border: cell.border = border
    if fmt:    cell.number_format = fmt

def hdr(cell, val):
    cell.value = val
    cell.font  = Font(name="Calibri", bold=False, size=9, color="FFFFFF")
    cell.fill  = BLACK_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = none_border

def section_title(ws, r, last_col, text):
    ws.row_dimensions[r].height = 14
    plain(ws.cell(r, 2), text, size=8, color="888888", border=bot_medium)
    for col in range(3, last_col + 1):
        ws.cell(r, col).border = bot_medium

ws.column_dimensions["A"].width = 1
ws.column_dimensions["B"].width = 9
for col in range(3, T + 4):
    ws.column_dimensions[get_column_letter(col)].width = 5.5
last_col = T + 2

# ── Title ──────────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 26
ws.merge_cells(f"B1:{get_column_letter(last_col)}1")
plain(ws.cell(1, 2), "Assignment 4b - Realized demand evaluation (permanent expansion)", bold=True, size=13)

# ── Cost summary ───────────────────────────────────────────────────────────
ws.row_dimensions[2].height = 4
cost_rows = [
    ("Setup cost",          setup_cost_fixed, False),
    ("Holding cost",        total_holding,    False),
    ("Backorder cost",      total_backorder,  False),
    ("Investment cost X",   invest_X,         False),
    ("Investment cost Y",   invest_Y,         False),
    ("Total cost (4b)",     total_cost_4b,    True),
    ("Total cost (4a)",     cost_4a,          False),
    ("Difference",          total_cost_4b - cost_4a, False),
]
for r, (label, val, bold) in enumerate(cost_rows, start=3):
    ws.row_dimensions[r].height = 16
    plain(ws.cell(r, 2), label, size=9, color="555555")
    fmt = '"+"€"#,##0.00;"-"€"#,##0.00' if label == "Difference" else '"€"#,##0.00'
    plain(ws.cell(r, 3), round(val, 2), bold=bold, size=9, fmt=fmt)
    ws.merge_cells(f"C{r}:{get_column_letter(last_col)}{r}")

# ── Service metrics ────────────────────────────────────────────────────────
ws.row_dimensions[11].height = 6
ws.row_dimensions[12].height = 14
section_title(ws, 12, last_col, "Service metrics (end product)")
for r, (label, text, red) in enumerate([
    ("Service level",   f"{service_level*100:.1f}%  ({periods_no_bo}/{T} periods without backorder)", False),
    ("Fill rate",       f"{fill_rate*100:.2f}%  ({total_realized-total_new_bo:,.0f} / {total_realized:,.0f} units on time)", False),
    ("Total backorder", f"{total_new_bo:,.0f} units", total_new_bo > 0),
], start=13):
    ws.row_dimensions[r].height = 16
    plain(ws.cell(r, 2), label, size=9, color="555555")
    plain(ws.cell(r, 3), text, size=9, color="CC0000" if red else "000000")
    ws.merge_cells(f"C{r}:{get_column_letter(last_col)}{r}")

# ══════════════════════════════════════════════════════════════════════════
# Comparison table 1b / 2b / 3b / 4b
# ══════════════════════════════════════════════════════════════════════════
r = 17
section_title(ws, r, last_col, "Comparison: 1b vs 2b vs 3b vs 4b")
r += 1
ws.row_dimensions[r].height = 16

for c, w in [(2,22),(3,14),(4,14),(5,14),(6,14)]:
    ws.column_dimensions[get_column_letter(c)].width = w

hdr(ws.cell(r, 2), "Metric")
for col, lbl in [(3,"1b"),(4,"2b"),(5,"3b"),(6,"4b")]:
    hdr(ws.cell(r, col), lbl)

cmp_rows = [
    ("Total cost (EUR)",    "total",  '"€"#,##0.00'),
    ("Setup cost (EUR)",    "setup",  '"€"#,##0.00'),
    ("Holding cost (EUR)",  "holding",'"€"#,##0.00'),
    ("Backorder cost (EUR)","bo",     '"€"#,##0.00'),
    ("Service level (%)",   "sl",     '0.0'),
    ("Fill rate (%)",       "fr",     '0.00'),
]
vals_4b = dict(total=total_cost_4b, setup=setup_cost_fixed,
               holding=total_holding, bo=total_backorder,
               sl=service_level*100, fr=fill_rate*100)

for label, key, fmt in cmp_rows:
    r += 1
    ws.row_dimensions[r].height = 15
    plain(ws.cell(r, 2), label, size=9, border=bot_thin)
    for col, src in [(3,"1b"),(4,"2b"),(5,"3b")]:
        v = refs.get(src, {}).get(key) if has_ref else None
        if v is not None:
            plain(ws.cell(r, col), round(float(v),2), size=9,
                  align="right", fmt=fmt, border=bot_thin)
        else:
            plain(ws.cell(r, col), "n/a", size=9, align="right",
                  color="AAAAAA", border=bot_thin)
    plain(ws.cell(r, 6), round(vals_4b[key], 2), bold=True,
          size=9, align="right", fmt=fmt, border=bot_thin)

# ══════════════════════════════════════════════════════════════════════════
# Backorder detail per period
# ══════════════════════════════════════════════════════════════════════════
# reset col widths for week table
ws.column_dimensions["B"].width = 9
for col in range(3, T + 4):
    ws.column_dimensions[get_column_letter(col)].width = 5.5

r += 2
section_title(ws, r, last_col, f"Backorder schedule - {END_PRODUCT} (end product only)")
r += 1
ws.row_dimensions[r].height = 16
hdr(ws.cell(r, 2), "Part")
for t in periods:
    hdr(ws.cell(r, t + 2), str(t))

r += 1
ws.row_dimensions[r].height = 15
plain(ws.cell(r, 2), END_PRODUCT, size=9, border=bot_thin)
for t in periods:
    v = round(BO[t].X)
    if v > 0:
        plain(ws.cell(r, t+2), v, bold=True, size=9, align="center",
              color="CC0000", fmt='#,##0', border=bot_thin)
    else:
        plain(ws.cell(r, t+2), "", size=9, align="center", border=bot_thin)

# ══════════════════════════════════════════════════════════════════════════
# Demand vs delivered
# ══════════════════════════════════════════════════════════════════════════
r += 2
section_title(ws, r, last_col, "Demand vs delivered (end product)")
r += 1
ws.row_dimensions[r].height = 16
hdr(ws.cell(r, 2), "")
for t in periods:
    hdr(ws.cell(r, t+2), str(t))

r += 1
ws.row_dimensions[r].height = 15
plain(ws.cell(r, 2), "Demand", size=9, border=bot_thin)
for t in periods:
    plain(ws.cell(r, t+2), DEMAND_REALIZED[t-1], size=9, align="center",
          fmt='#,##0', border=bot_thin)

r += 1
ws.row_dimensions[r].height = 15
plain(ws.cell(r, 2), "Delivered", size=9, border=bot_thin)
for t in periods:
    bo_prev = BO[t-1].X if t>1 else 0.0
    delivered = DEMAND_REALIZED[t-1] - max(0.0, new_bo[t-1])
    shortage  = delivered < DEMAND_REALIZED[t-1]
    plain(ws.cell(r, t+2), round(delivered), size=9, align="center",
          color="CC0000" if shortage else "000000",
          fmt='#,##0', border=bot_thin)

# ══════════════════════════════════════════════════════════════════════════
# Production schedule
# ══════════════════════════════════════════════════════════════════════════
r += 2
section_title(ws, r, last_col, "Production / order schedule (units, from 4a plan)")
r += 1
ws.row_dimensions[r].height = 16
hdr(ws.cell(r, 2), "Part")
for t in periods:
    hdr(ws.cell(r, t+2), str(t))

for i in PARTS:
    r += 1
    ws.row_dimensions[r].height = 15
    plain(ws.cell(r, 2), i, size=9, border=bot_thin)
    for t in periods:
        val = round(fixed_x[i][t]) if fixed_x[i][t] > 0.5 else ""
        plain(ws.cell(r, t+2), val, bold=bool(val), size=9, align="center",
              fmt='#,##0' if val != "" else None, border=bot_thin)

# ══════════════════════════════════════════════════════════════════════════
# Inventory levels (realized)
# ══════════════════════════════════════════════════════════════════════════
r += 2
section_title(ws, r, last_col, "Inventory levels (end of period, realized)")
r += 1
ws.row_dimensions[r].height = 16
hdr(ws.cell(r, 2), "Part")
for t in periods:
    hdr(ws.cell(r, t+2), str(t))

for i in PARTS:
    r += 1
    ws.row_dimensions[r].height = 15
    plain(ws.cell(r, 2), i, size=9, border=bot_thin)
    for t in periods:
        v = clean(I[i,t].X)
        plain(ws.cell(r, t+2), v, size=9, align="center",
              color="BBBBBB" if v == 0 else "000000",
              fmt='#,##0', border=bot_thin)

# ══════════════════════════════════════════════════════════════════════════
# Capacity utilisation
# ══════════════════════════════════════════════════════════════════════════
r += 2
section_title(ws, r, last_col,
              f"Capacity utilisation  (X: {round(cap_x_new)} u/wk  |  Y: {round(cap_y_new)} min/wk)")
r += 1
ws.row_dimensions[r].height = 16
hdr(ws.cell(r, 2), "")
for t in periods:
    hdr(ws.cell(r, t+2), str(t))

cap_rows = [
    ("X  (units)",  lambda t: sum(PROC_X.get(i,0)*fixed_x[i][t] for i in PARTS), cap_x_new, '#,##0'),
    ("Y  (min)",    lambda t: sum(PROC_Y.get(i,0)*fixed_x[i][t] for i in PARTS), cap_y_new, '#,##0'),
    ("X  (% used)", lambda t: sum(PROC_X.get(i,0)*fixed_x[i][t] for i in PARTS)/cap_x_new, 1, '0%'),
    ("Y  (% used)", lambda t: sum(PROC_Y.get(i,0)*fixed_x[i][t] for i in PARTS)/cap_y_new, 1, '0%'),
]
for label, fn, cap, fmt in cap_rows:
    r += 1
    ws.row_dimensions[r].height = 15
    plain(ws.cell(r, 2), label, size=9, border=bot_thin)
    for t in periods:
        raw = fn(t)
        over = raw > 0.95
        col  = "CC0000" if over else ("000000" if raw > 0 else "BBBBBB")
        plain(ws.cell(r, t+2), round(raw,3) if raw>0 else "", bold=over,
              size=9, align="center", color=col,
              fmt=fmt if raw>0 else None, border=bot_thin)

# ══════════════════════════════════════════════════════════════════════════
# Setup decisions
# ══════════════════════════════════════════════════════════════════════════
r += 2
section_title(ws, r, last_col, "Setup decisions")
r += 1
ws.row_dimensions[r].height = 16
hdr(ws.cell(r, 2), "Part")
for t in periods:
    hdr(ws.cell(r, t+2), str(t))

for i in PARTS:
    r += 1
    ws.row_dimensions[r].height = 15
    plain(ws.cell(r, 2), i, size=9, border=bot_thin)
    for t in periods:
        plain(ws.cell(r, t+2), "x" if fixed_x[i][t] > 0.5 else "",
              bold=True, size=9, align="center", border=bot_thin)

wb.save(OUTPUT_FILE)
print(f"Results written to {OUTPUT_FILE} -> sheet '{SHEET_NAME}'")
print(f"Total cost (4b): EUR {total_cost_4b:,.2f}")
print(f"  Setup:         EUR {setup_cost_fixed:,.2f}")
print(f"  Holding:       EUR {total_holding:,.2f}")
print(f"  Backorder:     EUR {total_backorder:,.2f}")
print(f"  Investment X:  EUR {invest_X:,.2f}")
print(f"  Investment Y:  EUR {invest_Y:,.2f}")
print(f"Service level:   {service_level*100:.1f}%")
print(f"Fill rate:       {fill_rate*100:.2f}%")