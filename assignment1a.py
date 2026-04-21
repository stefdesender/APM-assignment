"""
APM Project 2026 - Assignment 1a
MIP model for production planning (infinite capacity, forecasted demand)
"""

import gurobipy as gp
from gurobipy import GRB
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from input_data import (
    PARTS, END_PRODUCT, T, BOM, LEAD_TIME, MIN_LOT,
    INIT_INV, SETUP_COST, HOLDING_COST, DEMAND_FORECAST
)

def get_parents(part):
    parents = {}
    for parent, children in BOM.items():
        if part in children:
            parents[parent] = children[part]
    return parents

model = gp.Model("APM_Assignment1a")
periods = range(1, T + 1)
parts   = PARTS

x = model.addVars(parts, periods, name="x", lb=0)
y = model.addVars(parts, periods, name="y", vtype=GRB.BINARY)
I = model.addVars(parts, periods, name="I", lb=0)

BIG_M = {i: sum(DEMAND_FORECAST) * 25 for i in parts}

model.setObjective(
    gp.quicksum(
        SETUP_COST[i] * y[i, t] + HOLDING_COST[i] * I[i, t]
        for i in parts for t in periods
    ),
    GRB.MINIMIZE
)

for i in parts:
    for t in periods:
        inv_prev = INIT_INV[i] if t == 1 else I[i, t - 1]
        order_placement_period = t - LEAD_TIME[i]
        receipts = x[i, order_placement_period] if order_placement_period >= 1 else 0
        internal_demand = gp.quicksum(BOM[p][i] * x[p, t] for p in get_parents(i))
        external_demand = DEMAND_FORECAST[t - 1] if i == END_PRODUCT else 0
        model.addConstr(I[i, t] == inv_prev + receipts - internal_demand - external_demand,
                        name=f"inv_balance_{i}_{t}")
        model.addConstr(x[i, t] >= MIN_LOT[i] * y[i, t], name=f"min_lot_{i}_{t}")
        model.addConstr(x[i, t] <= BIG_M[i] * y[i, t],   name=f"bigM_{i}_{t}")

model.optimize()

if model.status == GRB.OPTIMAL:
    total_setup   = sum(SETUP_COST[i]   * y[i, t].X for i in parts for t in periods)
    total_holding = sum(HOLDING_COST[i] * I[i, t].X for i in parts for t in periods)

    OUTPUT_FILE = "OUTPUT.xlsx"
    SHEET_NAME  = "Output_1a"

    if os.path.exists(OUTPUT_FILE):
        wb = load_workbook(OUTPUT_FILE)
        if SHEET_NAME in wb.sheetnames:
            del wb[SHEET_NAME]
        ws = wb.create_sheet(SHEET_NAME)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME

    # ── Palette (neutral, minimal) ─────────────────────────────────────────
    # Charcoal header, light warm gray rows, white base
    CHARCOAL    = PatternFill("solid", start_color="2C2C2A", end_color="2C2C2A")  # near-black
    STONE       = PatternFill("solid", start_color="F1EFE8", end_color="F1EFE8")  # warm off-white
    WHITE       = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
    ACCENT      = PatternFill("solid", start_color="E6F1FB", end_color="E6F1FB")  # very pale blue for totals
    DOT_GREEN   = PatternFill("solid", start_color="EAF3DE", end_color="EAF3DE")  # soft green for setups

    thin   = Side(style="thin",   color="D3D1C7")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    NO_BORDER = Border()

    def h(cell, text, small=False):
        """Header cell: charcoal bg, white text."""
        cell.value     = text
        cell.font      = Font(name="Arial", bold=False, color="FFFFFF",
                              size=8 if small else 10)
        cell.fill      = CHARCOAL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = BORDER

    def d(cell, val, bold=False, align="right", fill=WHITE, fmt=None, color="2C2C2A"):
        cell.value     = val
        cell.font      = Font(name="Arial", bold=bold, color=color, size=9)
        cell.fill      = fill
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border    = BORDER
        if fmt:
            cell.number_format = fmt

    def section_label(ws, row, col_end, text):
        """Small all-caps section divider."""
        ws.row_dimensions[row].height = 18
        ws.merge_cells(f"B{row}:{get_column_letter(col_end)}{row}")
        c = ws.cell(row, 2, text)
        c.font      = Font(name="Arial", bold=True, color="888780", size=8)
        c.fill      = WHITE
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = NO_BORDER

    # ── Column layout ──────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 1.5   # left margin
    ws.column_dimensions["B"].width = 10    # part label
    for col in range(3, T + 4):
        ws.column_dimensions[get_column_letter(col)].width = 5.5

    last_col = T + 2   # column index of last week

    # ══════════════════════════════════════════════════════════════════════
    # Row 1 – Title
    # ══════════════════════════════════════════════════════════════════════
    ws.row_dimensions[1].height = 28
    ws.merge_cells(f"B1:{get_column_letter(last_col)}1")
    tc = ws.cell(1, 2, "Assignment 1a  —  Optimal solution")
    tc.font      = Font(name="Arial", bold=False, size=13, color="2C2C2A")
    tc.fill      = WHITE
    tc.alignment = Alignment(horizontal="left", vertical="center")
    tc.border    = NO_BORDER

    # ══════════════════════════════════════════════════════════════════════
    # Rows 3-5 – Cost summary (3 metric cells, no table)
    # ══════════════════════════════════════════════════════════════════════
    ws.row_dimensions[2].height = 6
    for r, (label, val) in enumerate([
        ("Total cost",   model.ObjVal),
        ("Setup cost",   total_setup),
        ("Holding cost", total_holding),
    ], start=3):
        ws.row_dimensions[r].height = 18
        lc = ws.cell(r, 2, label)
        lc.font      = Font(name="Arial", size=8, color="888780")
        lc.fill      = WHITE
        lc.alignment = Alignment(horizontal="left", vertical="center")
        lc.border    = NO_BORDER

        vc = ws.cell(r, 3, round(val, 2))
        vc.font         = Font(name="Arial", size=9,
                               bold=(r == 3), color="2C2C2A")
        vc.fill         = ACCENT if r == 3 else WHITE
        vc.alignment    = Alignment(horizontal="left", vertical="center")
        vc.number_format = '€#,##0.00'
        vc.border       = NO_BORDER
        ws.merge_cells(f"C{r}:{get_column_letter(last_col)}{r}")

    # ══════════════════════════════════════════════════════════════════════
    # Production schedule
    # ══════════════════════════════════════════════════════════════════════
    r = 7
    section_label(ws, r, last_col, "PRODUCTION / ORDER SCHEDULE (units)")

    r += 1
    ws.row_dimensions[r].height = 16
    h(ws.cell(r, 2), "Part")
    for t in periods:
        h(ws.cell(r, t + 2), str(t), small=True)

    for idx, i in enumerate(parts, start=1):
        r += 1
        ws.row_dimensions[r].height = 15
        alt = STONE if idx % 2 == 0 else WHITE
        d(ws.cell(r, 2), i, bold=False, align="left", fill=alt, color="444441")
        for t in periods:
            val = round(x[i, t].X) if x[i, t].X > 0.5 else ""
            d(ws.cell(r, t + 2), val, align="center", fill=alt,
              fmt='#,##0' if val != "" else None,
              bold=bool(val), color="2C2C2A" if val else "888780")

    # ══════════════════════════════════════════════════════════════════════
    # Inventory levels
    # ══════════════════════════════════════════════════════════════════════
    r += 2
    section_label(ws, r, last_col, "INVENTORY LEVELS (end of period)")

    r += 1
    ws.row_dimensions[r].height = 16
    h(ws.cell(r, 2), "Part")
    for t in periods:
        h(ws.cell(r, t + 2), str(t), small=True)

    for idx, i in enumerate(parts, start=1):
        r += 1
        ws.row_dimensions[r].height = 15
        alt = STONE if idx % 2 == 0 else WHITE
        d(ws.cell(r, 2), i, bold=False, align="left", fill=alt, color="444441")
        for t in periods:
            v = round(I[i, t].X)
            d(ws.cell(r, t + 2), v, align="center", fill=alt, fmt='#,##0',
              color="2C2C2A" if v > 0 else "B4B2A9")

    # ══════════════════════════════════════════════════════════════════════
    # Setup decisions
    # ══════════════════════════════════════════════════════════════════════
    r += 2
    section_label(ws, r, last_col, "SETUP DECISIONS")

    r += 1
    ws.row_dimensions[r].height = 16
    h(ws.cell(r, 2), "Part")
    for t in periods:
        h(ws.cell(r, t + 2), str(t), small=True)

    for idx, i in enumerate(parts, start=1):
        r += 1
        ws.row_dimensions[r].height = 15
        alt = STONE if idx % 2 == 0 else WHITE
        d(ws.cell(r, 2), i, bold=False, align="left", fill=alt, color="444441")
        for t in periods:
            setup = int(round(y[i, t].X))
            vc = ws.cell(r, t + 2)
            if setup:
                d(vc, "●", align="center", fill=DOT_GREEN, color="3B6D11")
            else:
                d(vc, "", align="center", fill=alt)

    wb.save(OUTPUT_FILE)
    print(f"\nResults written to {OUTPUT_FILE} → sheet '{SHEET_NAME}'")
    print(f"Total cost:    €{model.ObjVal:,.2f}")
    print(f"  Setup cost:  €{total_setup:,.2f}")
    print(f"  Holding cost:€{total_holding:,.2f}")

else:
    print(f"Model status: {model.status} — no optimal solution found.")