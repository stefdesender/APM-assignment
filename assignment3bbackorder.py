"""
APM Project 2026 - Assignment 3b
"""

import gurobipy as gp
from gurobipy import GRB
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from input_data import (
    PARTS, END_PRODUCT, T, BOM, LEAD_TIME, MIN_LOT,
    INIT_INV, SETUP_COST, HOLDING_COST, DEMAND_FORECAST,
    DEMAND_REALIZED, BACKORDER_COST
)

def get_parents(part):
    parents = {}
    for parent, children in BOM.items():
        if part in children:
            parents[parent] = children[part]
    return parents

def clean_num(val, tol=1e-6):
    v = float(val)
    return 0.0 if abs(v) < tol else round(v, 6)

# ── Re-solve 3a to get fixed plan ─────────────────────────────────────────
CAPACITY_X        = 800
CAPACITY_X_OT_MAX = 300
COST_OT_X         = 2
CAPACITY_Y        = 7 * 24 * 60 - 80
CAPACITY_Y_OT_MAX = 38
COST_OT_Y         = 120
PROC_TIME_Y       = {'B1401': 3, 'B2302': 2}

periods = range(1, T + 1)

model = gp.Model("APM_3a_reload")
model.setParam("OutputFlag", 0)
x    = model.addVars(PARTS, periods, name="x", vtype=GRB.INTEGER, lb=0)
y    = model.addVars(PARTS, periods, name="y", vtype=GRB.BINARY)
I    = model.addVars(PARTS, periods, name="I", vtype=GRB.INTEGER, lb=0)
ot_x = model.addVars(periods, name="ot_x", vtype=GRB.INTEGER, lb=0, ub=CAPACITY_X_OT_MAX)
ot_y = model.addVars(periods, name="ot_y", lb=0, ub=CAPACITY_Y_OT_MAX)
BIG_M = {i: sum(DEMAND_FORECAST) * 25 for i in PARTS}

model.setObjective(
    gp.quicksum(SETUP_COST[i]*y[i,t] + HOLDING_COST[i]*I[i,t]
                for i in PARTS for t in periods)
    + gp.quicksum(COST_OT_X*ot_x[t] + COST_OT_Y*ot_y[t] for t in periods),
    GRB.MINIMIZE)

for i in PARTS:
    for t in periods:
        inv_prev = INIT_INV[i] if t == 1 else I[i, t-1]
        op = t - LEAD_TIME[i]
        receipts = x[i, op] if op >= 1 else 0
        int_dem  = gp.quicksum(BOM[p][i]*x[p,t] for p in get_parents(i))
        ext_dem  = DEMAND_FORECAST[t-1] if i == END_PRODUCT else 0
        model.addConstr(I[i,t] == inv_prev + receipts - int_dem - ext_dem)
        model.addConstr(x[i,t] >= MIN_LOT[i]*y[i,t])
        model.addConstr(x[i,t] <= BIG_M[i]*y[i,t])

for t in periods:
    model.addConstr(x['E2801',t] <= CAPACITY_X + ot_x[t])
    model.addConstr(PROC_TIME_Y['B1401']*x['B1401',t] + PROC_TIME_Y['B2302']*x['B2302',t]
                    <= CAPACITY_Y + 60*ot_y[t])

model.optimize()

schedule       = {i: {t: x[i,t].X for t in periods} for i in PARTS}
setup_cost_3a  = sum(SETUP_COST[i]*y[i,t].X for i in PARTS for t in periods)
total_ot_x_3a  = sum(COST_OT_X * ot_x[t].X for t in periods)
total_ot_y_3a  = sum(COST_OT_Y * ot_y[t].X for t in periods)
total_ot_3a    = total_ot_x_3a + total_ot_y_3a
cost_3a        = model.ObjVal

# overtime per week for output
ot_x_vals = {t: clean_num(ot_x[t].X) for t in periods}
ot_y_vals = {t: clean_num(ot_y[t].X) for t in periods}

# ── Simulate with realized demand ─────────────────────────────────────────
inv   = {i: {} for i in PARTS}
h_pos = {i: {} for i in PARTS}
net_end  = {}
bo       = {}
held_end = {}

for i in PARTS:
    if i == END_PRODUCT:
        continue
    inv_prev = float(INIT_INV[i])
    for t in periods:
        op = t - LEAD_TIME[i]
        receipts = schedule[i].get(op, 0.0) if op >= 1 else 0.0
        int_dem  = sum(qty * schedule[p].get(t, 0.0) for p, qty in get_parents(i).items())
        net = inv_prev + receipts - int_dem
        inv[i][t]   = clean_num(net)
        h_pos[i][t] = clean_num(max(0.0, net))
        inv_prev = net

inv_prev = float(INIT_INV[END_PRODUCT])
bo_prev  = 0.0
for t in periods:
    op = t - LEAD_TIME[END_PRODUCT]
    receipts = schedule[END_PRODUCT].get(op, 0.0) if op >= 1 else 0.0
    net = inv_prev - bo_prev + receipts - DEMAND_REALIZED[t-1]
    net_end[t]   = clean_num(net)
    bo[t]        = clean_num(max(0.0, -net))
    held_end[t]  = clean_num(max(0.0, net))
    inv[END_PRODUCT][t]   = held_end[t]
    h_pos[END_PRODUCT][t] = held_end[t]
    inv_prev = held_end[t]
    bo_prev  = bo[t]

total_holding   = sum(HOLDING_COST[i] * h_pos[i][t] for i in PARTS for t in periods)
total_backorder = sum(BACKORDER_COST * bo[t] for t in periods)
total_cost_3b   = setup_cost_3a + total_holding + total_ot_3a + total_backorder

periods_no_bo = sum(1 for t in periods if bo[t] == 0)
service_level = periods_no_bo / T
total_demand  = sum(DEMAND_REALIZED)
total_new_bo  = 0.0
prev_bo = 0.0
for t in periods:
    new_bo = max(0.0, bo[t] - prev_bo)
    total_new_bo += new_bo
    prev_bo = bo[t]
fill_rate    = 1.0 - (total_new_bo / total_demand) if total_demand > 0 else 1.0
units_on_time = total_demand - total_new_bo

# ── Read 1b and 2b results from OUTPUT.xlsx for comparison ────────────────
def read_cost_cell(ws, row):
    v = ws.cell(row, 3).value
    try:
        return float(v)
    except Exception:
        return None

try:
    wb_ref = load_workbook("OUTPUT.xlsx", data_only=True)
    # 1b: rows 3=setup,4=holding,5=bo,6=total,  sl row11, fr row12
    ws1b = wb_ref["Output_1b"]
    c1b  = {
        "setup":   read_cost_cell(ws1b, 3),
        "holding": read_cost_cell(ws1b, 4),
        "bo":      read_cost_cell(ws1b, 5),
        "total":   read_cost_cell(ws1b, 6),
    }
    import re
    def parse_pct(s):
        m = re.search(r"[\d.]+", str(s) if s else "")
        return float(m.group()) if m else None
    c1b["sl"] = parse_pct(ws1b.cell(11,3).value)
    c1b["fr"] = parse_pct(ws1b.cell(12,3).value)

    # 2b: rows 3=setup,4=holding,5=bo,6=total, sl row9, fr row10
    ws2b = wb_ref["Output_2b"]
    c2b  = {
        "setup":   read_cost_cell(ws2b, 3),
        "holding": read_cost_cell(ws2b, 4),
        "bo":      read_cost_cell(ws2b, 5),
        "total":   read_cost_cell(ws2b, 6),
        "sl":      parse_pct(ws2b.cell(9,3).value),
        "fr":      parse_pct(ws2b.cell(10,3).value),
    }
    has_ref = True
except Exception:
    has_ref = False

# ── Write to OUTPUT.xlsx ───────────────────────────────────────────────────
OUTPUT_FILE = "OUTPUT.xlsx"
SHEET_NAME  = "Output_3b"

if os.path.exists(OUTPUT_FILE):
    wb = load_workbook(OUTPUT_FILE)
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)
else:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

# ── Styles ────────────────────────────────────────────────────────────────
NO_FILL    = PatternFill(fill_type=None)
BLACK_FILL = PatternFill("solid", start_color="000000", end_color="000000")
none_border = Border()
bot_medium  = Border(bottom=Side(style="medium", color="000000"))
bot_thin    = Border(bottom=Side(style="thin",   color="CCCCCC"))

def plain(cell, val, bold=False, align="left", size=9, color="000000", fmt=None, border=None):
    cell.value = val
    cell.font  = Font(name="Calibri", bold=bold, size=size, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.fill  = NO_FILL
    if border: cell.border = border
    if fmt:    cell.number_format = fmt

def hdr(cell, val):
    cell.value = val
    cell.font  = Font(name="Calibri", bold=False, size=9, color="FFFFFF")
    cell.fill  = BLACK_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = none_border

def section_title(ws, r, last_col, text):
    ws.row_dimensions[r].height = 14
    plain(ws.cell(r, 2), text, size=8, color="888888", border=bot_medium)
    for col in range(3, last_col + 1):
        ws.cell(r, col).border = bot_medium

ws.column_dimensions["A"].width = 1
ws.column_dimensions["B"].width = 9
for col in range(3, T + 4):
    ws.column_dimensions[get_column_letter(col)].width = 5.5
last_col = T + 2

# ── Title ──────────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 26
ws.merge_cells(f"B1:{get_column_letter(last_col)}1")
plain(ws.cell(1, 2), "Assignment 3b - Realized demand evaluation (finite capacity + overtime)", bold=True, size=13)

# ── Cost summary ───────────────────────────────────────────────────────────
ws.row_dimensions[2].height = 4
cost_rows = [
    ("Setup cost",          setup_cost_3a,   False),
    ("Holding cost",        total_holding,   False),
    ("Overtime cost X",     total_ot_x_3a,   False),
    ("Overtime cost Y",     total_ot_y_3a,   False),
    ("Backorder cost",      total_backorder, False),
    ("Total cost",          total_cost_3b,   True),
    ("Total cost (3a)",     cost_3a,         False),
    ("Difference",          total_cost_3b - cost_3a, False),
]
for r, (label, val, bold) in enumerate(cost_rows, start=3):
    ws.row_dimensions[r].height = 16
    plain(ws.cell(r, 2), label, size=9, color="555555")
    fmt = '"+"€"#,##0.00;"-"€"#,##0.00' if label == "Difference" else '"€"#,##0.00'
    plain(ws.cell(r, 3), round(clean_num(val), 2), bold=bold, size=9, fmt=fmt)
    ws.merge_cells(f"C{r}:{get_column_letter(last_col)}{r}")

# ── Service metrics ────────────────────────────────────────────────────────
ws.row_dimensions[11].height = 6
ws.row_dimensions[12].height = 14
section_title(ws, 12, last_col, "Service metrics (end product)")
for r, (label, text) in enumerate([
    ("Service level", f"{service_level*100:.1f}%  ({periods_no_bo}/{T} periods without backorder)"),
    ("Fill rate",     f"{fill_rate*100:.2f}%  ({units_on_time:,.0f} / {total_demand:,.0f} units on time)"),
    ("Total backorder", f"{sum(bo.values()):,.0f} units"),
], start=13):
    ws.row_dimensions[r].height = 16
    plain(ws.cell(r, 2), label, size=9, color="555555")
    col = "CC0000" if (label == "Total backorder" and sum(bo.values()) > 0) else "000000"
    plain(ws.cell(r, 3), text, size=9, color=col)
    ws.merge_cells(f"C{r}:{get_column_letter(last_col)}{r}")

# ══════════════════════════════════════════════════════════════════════════
# Comparison table 1b / 2b / 3b
# ══════════════════════════════════════════════════════════════════════════
r = 17
section_title(ws, r, last_col, "Comparison: 1b vs 2b vs 3b")
r += 1
ws.row_dimensions[r].height = 16

# fixed columns for comparison table
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 14

hdr(ws.cell(r, 2), "Metric")
hdr(ws.cell(r, 3), "1b")
hdr(ws.cell(r, 4), "2b")
hdr(ws.cell(r, 5), "3b")

cmp_rows = [
    ("Total cost (EUR)",     c1b.get("total")   if has_ref else None, c2b.get("total")   if has_ref else None, total_cost_3b,   '"€"#,##0.00'),
    ("Setup cost (EUR)",     c1b.get("setup")   if has_ref else None, c2b.get("setup")   if has_ref else None, setup_cost_3a,   '"€"#,##0.00'),
    ("Holding cost (EUR)",   c1b.get("holding") if has_ref else None, c2b.get("holding") if has_ref else None, total_holding,   '"€"#,##0.00'),
    ("Backorder cost (EUR)", c1b.get("bo")      if has_ref else None, c2b.get("bo")      if has_ref else None, total_backorder, '"€"#,##0.00'),
    ("Service level (%)",    c1b.get("sl")      if has_ref else None, c2b.get("sl")      if has_ref else None, service_level*100, '0.0'),
    ("Fill rate (%)",        c1b.get("fr")      if has_ref else None, c2b.get("fr")      if has_ref else None, fill_rate*100,   '0.00'),
]

for idx, (label, v1b, v2b, v3b, fmt) in enumerate(cmp_rows, start=1):
    r += 1
    ws.row_dimensions[r].height = 15
    plain(ws.cell(r, 2), label, size=9, border=bot_thin)
    for col, val in [(3, v1b), (4, v2b), (5, v3b)]:
        if val is not None:
            plain(ws.cell(r, col), round(float(val), 2), size=9, align="right",
                  fmt=fmt, border=bot_thin,
                  bold=(col == 5))
        else:
            plain(ws.cell(r, col), "n/a", size=9, align="right",
                  color="AAAAAA", border=bot_thin)

# ══════════════════════════════════════════════════════════════════════════
# Production schedule
# ══════════════════════════════════════════════════════════════════════════
# reset col widths for week table
ws.column_dimensions["B"].width = 9
for col in range(3, T + 4):
    ws.column_dimensions[get_column_letter(col)].width = 5.5

r += 2
section_title(ws, r, last_col, "Production / order schedule (units, from 3a plan)")
r += 1
ws.row_dimensions[r].height = 16
hdr(ws.cell(r, 2), "Part")
for t in periods:
    hdr(ws.cell(r, t + 2), str(t))

for i in PARTS:
    r += 1
    ws.row_dimensions[r].height = 15
    plain(ws.cell(r, 2), i, size=9, border=bot_thin)
    for t in periods:
        val = round(schedule[i][t]) if schedule[i][t] > 0.5 else ""
        plain(ws.cell(r, t + 2), val, bold=bool(val), size=9, align="center",
              fmt='#,##0' if val != "" else None, border=bot_thin)

# ══════════════════════════════════════════════════════════════════════════
# Inventory levels (realized)
# ══════════════════════════════════════════════════════════════════════════
r += 2
section_title(ws, r, last_col, "Inventory levels (end of period, realized)")
r += 1
ws.row_dimensions[r].height = 16
hdr(ws.cell(r, 2), "Part")
for t in periods:
    hdr(ws.cell(r, t + 2), str(t))

for i in PARTS:
    r += 1
    ws.row_dimensions[r].height = 15
    plain(ws.cell(r, 2), i, size=9, border=bot_thin)
    for t in periods:
        v = round(inv[i][t])
        col = "CC0000" if v < 0 else ("BBBBBB" if v == 0 else "000000")
        plain(ws.cell(r, t + 2), v, size=9, align="center",
              color=col, fmt='#,##0;-#,##0', border=bot_thin)

# ══════════════════════════════════════════════════════════════════════════
# Backorder schedule
# ══════════════════════════════════════════════════════════════════════════
r += 2
section_title(ws, r, last_col, f"Backorder schedule - {END_PRODUCT} (end product only)")
r += 1
ws.row_dimensions[r].height = 16
hdr(ws.cell(r, 2), "Part")
for t in periods:
    hdr(ws.cell(r, t + 2), str(t))

r += 1
ws.row_dimensions[r].height = 15
plain(ws.cell(r, 2), END_PRODUCT, size=9, border=bot_thin)
for t in periods:
    v = round(bo[t])
    if v > 0:
        plain(ws.cell(r, t + 2), v, bold=True, size=9, align="center",
              color="CC0000", fmt='#,##0', border=bot_thin)
    else:
        plain(ws.cell(r, t + 2), "", size=9, align="center", border=bot_thin)

# ══════════════════════════════════════════════════════════════════════════
# Overtime usage (from 3a plan)
# ══════════════════════════════════════════════════════════════════════════
r += 2
section_title(ws, r, last_col, "Overtime usage (from 3a plan, unchanged)")
r += 1
ws.row_dimensions[r].height = 16
hdr(ws.cell(r, 2), "")
for t in periods:
    hdr(ws.cell(r, t + 2), str(t))

r += 1
ws.row_dimensions[r].height = 15
plain(ws.cell(r, 2), "X  (units OT)", size=9, border=bot_thin)
for t in periods:
    v = round(ot_x_vals[t])
    col = "CC0000" if v > 0 else "BBBBBB"
    plain(ws.cell(r, t + 2), v if v else "", bold=(v > 0), size=9, align="center",
          color=col, fmt='#,##0' if v else None, border=bot_thin)

r += 1
ws.row_dimensions[r].height = 15
plain(ws.cell(r, 2), "Y  (hours OT)", size=9, border=bot_thin)
for t in periods:
    v = round(ot_y_vals[t], 2)
    col = "CC0000" if v > 0 else "BBBBBB"
    plain(ws.cell(r, t + 2), v if v else "", bold=(v > 0), size=9, align="center",
          color=col, fmt='0.00' if v else None, border=bot_thin)

# ══════════════════════════════════════════════════════════════════════════
# Setup decisions
# ══════════════════════════════════════════════════════════════════════════
r += 2
section_title(ws, r, last_col, "Setup decisions")
r += 1
ws.row_dimensions[r].height = 16
hdr(ws.cell(r, 2), "Part")
for t in periods:
    hdr(ws.cell(r, t + 2), str(t))

for i in PARTS:
    r += 1
    ws.row_dimensions[r].height = 15
    plain(ws.cell(r, 2), i, size=9, border=bot_thin)
    for t in periods:
        setup = 1 if schedule[i][t] > 0.5 else 0
        plain(ws.cell(r, t + 2), "x" if setup else "",
              bold=True, size=9, align="center", border=bot_thin)

wb.save(OUTPUT_FILE)
print(f"Results written to {OUTPUT_FILE} -> sheet '{SHEET_NAME}'")
print(f"Total cost (3b): EUR {clean_num(total_cost_3b):,.2f}")
print(f"  Setup:         EUR {clean_num(setup_cost_3a):,.2f}")
print(f"  Holding:       EUR {clean_num(total_holding):,.2f}")
print(f"  Overtime X:    EUR {clean_num(total_ot_x_3a):,.2f}")
print(f"  Overtime Y:    EUR {clean_num(total_ot_y_3a):,.2f}")
print(f"  Backorder:     EUR {clean_num(total_backorder):,.2f}")
print(f"Service level:   {service_level*100:.1f}%")
print(f"Fill rate:       {fill_rate*100:.2f}%")