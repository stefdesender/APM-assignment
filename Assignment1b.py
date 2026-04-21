"""
APM Project 2026 - Assignment 1b
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from input_data import (
    PARTS, END_PRODUCT, T, BOM, LEAD_TIME,
    INIT_INV, SETUP_COST, HOLDING_COST,
    DEMAND_REALIZED, BACKORDER_COST,
)

# ── Load 1a plan from OUTPUT.xlsx ─────────────────────────────────────────
# We re-read the production schedule from the Output_1a sheet
from openpyxl import load_workbook as _lw

_wb = _lw("OUTPUT.xlsx", data_only=True)
_ws = _wb["Output_1a"]

# Cost summary is in rows 3-5, col C
total_cost_1a  = _ws.cell(3, 3).value
total_setup_1a = _ws.cell(4, 3).value

# Re-run 1a model to get actual variable values
# (easier to just re-solve than parse the sheet)
import gurobipy as gp
from gurobipy import GRB

def get_parents(part):
    parents = {}
    for parent, children in BOM.items():
        if part in children:
            parents[parent] = children[part]
    return parents

from input_data import DEMAND_FORECAST, MIN_LOT

periods = range(1, T + 1)

model = gp.Model("APM_1a_reload")
model.setParam("OutputFlag", 0)
x = model.addVars(PARTS, periods, name="x", lb=0)
y = model.addVars(PARTS, periods, name="y", vtype=GRB.BINARY)
I = model.addVars(PARTS, periods, name="I", lb=0)
BIG_M = {i: sum(DEMAND_FORECAST) * 25 for i in PARTS}

model.setObjective(
    gp.quicksum(SETUP_COST[i] * y[i,t] + HOLDING_COST[i] * I[i,t]
                for i in PARTS for t in periods), GRB.MINIMIZE)

for i in PARTS:
    for t in periods:
        inv_prev = INIT_INV[i] if t == 1 else I[i, t-1]
        op = t - LEAD_TIME[i]
        receipts = x[i, op] if op >= 1 else 0
        int_dem = gp.quicksum(BOM[p][i] * x[p,t] for p in get_parents(i))
        ext_dem = DEMAND_FORECAST[t-1] if i == END_PRODUCT else 0
        model.addConstr(I[i,t] == inv_prev + receipts - int_dem - ext_dem)
        model.addConstr(x[i,t] >= MIN_LOT[i] * y[i,t])
        model.addConstr(x[i,t] <= BIG_M[i] * y[i,t])

model.optimize()

schedule = {i: {t: x[i,t].X for t in periods} for i in PARTS}
cost_1a  = model.ObjVal

# ── Simulate 1b with realized demand ──────────────────────────────────────
inv   = {i: {} for i in PARTS}
bo    = {i: {} for i in PARTS}
h_pos = {i: {} for i in PARTS}

for i in PARTS:
    inv_prev = INIT_INV[i]
    for t in periods:
        op = t - LEAD_TIME[i]
        receipts = schedule[i].get(op, 0.0) if op >= 1 else 0.0
        internal_demand = sum(
            qty * schedule[p].get(t, 0.0)
            for p, qty in get_parents(i).items()
        )
        external_demand = DEMAND_REALIZED[t-1] if i == END_PRODUCT else 0.0
        net = inv_prev + receipts - internal_demand - external_demand
        inv[i][t]   = net
        bo[i][t]    = max(0.0, -net)
        h_pos[i][t] = max(0.0, net)
        inv_prev    = net

total_setup     = cost_1a - sum(HOLDING_COST[i] * I[i,t].X for i in PARTS for t in periods)
# use setup from 1a directly
total_setup     = sum(SETUP_COST[i] * y[i,t].X for i in PARTS for t in periods)
total_holding   = sum(HOLDING_COST[i] * h_pos[i][t] for i in PARTS for t in periods)
total_backorder = sum(BACKORDER_COST * bo[END_PRODUCT][t] for t in periods)
total_cost_1b   = total_setup + total_holding + total_backorder

periods_no_bo = sum(1 for t in periods if bo[END_PRODUCT][t] == 0)
service_level = periods_no_bo / T

total_demand = sum(DEMAND_REALIZED)
total_new_bo = 0.0
prev_bo = 0.0
for t in periods:
    new_bo = max(0.0, bo[END_PRODUCT][t] - prev_bo)
    total_new_bo += new_bo
    prev_bo = bo[END_PRODUCT][t]
fill_rate = 1.0 - (total_new_bo / total_demand) if total_demand > 0 else 1.0

# ── Write to OUTPUT.xlsx, sheet Output_1b ─────────────────────────────────
OUTPUT_FILE = "OUTPUT.xlsx"
SHEET_NAME  = "Output_1b"

if os.path.exists(OUTPUT_FILE):
    wb = load_workbook(OUTPUT_FILE)
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)
else:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

# ── Styles (identical to 1a) ───────────────────────────────────────────────
NO_FILL    = PatternFill(fill_type=None)
BLACK_FILL = PatternFill("solid", start_color="000000", end_color="000000")
RED_FILL   = PatternFill("solid", start_color="000000", end_color="000000")

none_border  = Border()
bot_medium   = Border(bottom=Side(style="medium", color="000000"))
bot_thin     = Border(bottom=Side(style="thin",   color="CCCCCC"))

def plain(cell, val, bold=False, align="left", size=9, color="000000", fmt=None, border=None):
    cell.value = val
    cell.font  = Font(name="Calibri", bold=bold, size=size, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.fill  = NO_FILL
    if border:
        cell.border = border
    if fmt:
        cell.number_format = fmt

def hdr(cell, val):
    cell.value = val
    cell.font  = Font(name="Calibri", bold=False, size=9, color="FFFFFF")
    cell.fill  = BLACK_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = none_border

def section_title(ws, r, last_col, text):
    ws.row_dimensions[r].height = 14
    plain(ws.cell(r, 2), text, size=8, color="888888", border=bot_medium)
    for col in range(3, last_col + 1):
        ws.cell(r, col).border = bot_medium

# ── Column widths ──────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 1
ws.column_dimensions["B"].width = 9
for col in range(3, T + 4):
    ws.column_dimensions[get_column_letter(col)].width = 5.5

last_col = T + 2

# ══════════════════════════════════════════════════════════════════════════
# Row 1: Title
# ══════════════════════════════════════════════════════════════════════════
ws.row_dimensions[1].height = 26
ws.merge_cells(f"B1:{get_column_letter(last_col)}1")
plain(ws.cell(1, 2), "Assignment 1b - Realized demand evaluation", bold=True, size=13)

# ══════════════════════════════════════════════════════════════════════════
# Rows 3-8: Cost summary + comparison
# ══════════════════════════════════════════════════════════════════════════
ws.row_dimensions[2].height = 4
summary_rows = [
    ("Setup cost",      total_setup,     '"€"#,##0.00', False),
    ("Holding cost",    total_holding,   '"€"#,##0.00', False),
    ("Backorder cost",  total_backorder, '"€"#,##0.00', False),
    ("Total cost (1b)", total_cost_1b,   '"€"#,##0.00', True),
    ("Total cost (1a)", cost_1a,         '"€"#,##0.00', False),
    ("Difference",      total_cost_1b - cost_1a, '"€"+#,##0.00;"€"-#,##0.00', False),
]
for r, (label, val, fmt, bold) in enumerate(summary_rows, start=3):
    ws.row_dimensions[r].height = 16
    plain(ws.cell(r, 2), label, size=9, color="555555")
    vc = ws.cell(r, 3)
    plain(vc, round(val, 2), bold=bold, size=9, fmt=fmt)
    ws.merge_cells(f"C{r}:{get_column_letter(last_col)}{r}")

# ══════════════════════════════════════════════════════════════════════════
# Rows 10-11: Service metrics
# ══════════════════════════════════════════════════════════════════════════
ws.row_dimensions[9].height = 6
section_title(ws, 10, last_col, "Service metrics (end product)")
ws.row_dimensions[11].height = 16
ws.row_dimensions[12].height = 16
plain(ws.cell(11, 2), "Service level", size=9, color="555555")
plain(ws.cell(11, 3), f"{service_level*100:.1f}%  ({periods_no_bo}/{T} periods without backorder)", size=9)
ws.merge_cells(f"C11:{get_column_letter(last_col)}11")
plain(ws.cell(12, 2), "Fill rate", size=9, color="555555")
plain(ws.cell(12, 3),
      f"{fill_rate*100:.2f}%  ({total_demand - total_new_bo:,.0f} / {total_demand:,.0f} units on time)",
      size=9)
ws.merge_cells(f"C12:{get_column_letter(last_col)}12")

# ══════════════════════════════════════════════════════════════════════════
# Production schedule
# ══════════════════════════════════════════════════════════════════════════
r = 14
section_title(ws, r, last_col, "Production / order schedule (units)")
r += 1
ws.row_dimensions[r].height = 16
hdr(ws.cell(r, 2), "Part")
for t in periods:
    hdr(ws.cell(r, t + 2), str(t))

for idx, i in enumerate(PARTS, start=1):
    r += 1
    ws.row_dimensions[r].height = 15
    plain(ws.cell(r, 2), i, size=9, border=bot_thin)
    for t in periods:
        val = round(schedule[i][t]) if schedule[i][t] > 0.5 else ""
        plain(ws.cell(r, t + 2), val, bold=bool(val), size=9, align="center",
              fmt='#,##0' if val != "" else None, border=bot_thin)

# ══════════════════════════════════════════════════════════════════════════
# Inventory levels
# ══════════════════════════════════════════════════════════════════════════
r += 2
section_title(ws, r, last_col, "Inventory levels (end of period, realized)")
r += 1
ws.row_dimensions[r].height = 16
hdr(ws.cell(r, 2), "Part")
for t in periods:
    hdr(ws.cell(r, t + 2), str(t))

for idx, i in enumerate(PARTS, start=1):
    r += 1
    ws.row_dimensions[r].height = 15
    plain(ws.cell(r, 2), i, size=9, border=bot_thin)
    for t in periods:
        v = round(inv[i][t])
        col = "000000" if v > 0 else ("CC0000" if v < 0 else "BBBBBB")
        plain(ws.cell(r, t + 2), v, size=9, align="center",
              color=col, fmt='#,##0;-#,##0', border=bot_thin)

# ══════════════════════════════════════════════════════════════════════════
# Backorder schedule (end product only)
# ══════════════════════════════════════════════════════════════════════════
r += 2
section_title(ws, r, last_col, f"Backorder schedule - {END_PRODUCT} (end product only)")
r += 1
ws.row_dimensions[r].height = 16
hdr(ws.cell(r, 2), "Part")
for t in periods:
    hdr(ws.cell(r, t + 2), str(t))

r += 1
ws.row_dimensions[r].height = 15
plain(ws.cell(r, 2), END_PRODUCT, size=9, border=bot_thin)
for t in periods:
    v = round(bo[END_PRODUCT][t])
    if v > 0:
        plain(ws.cell(r, t + 2), v, bold=True, size=9, align="center",
              color="CC0000", fmt='#,##0', border=bot_thin)
    else:
        plain(ws.cell(r, t + 2), "", size=9, align="center", border=bot_thin)

# ══════════════════════════════════════════════════════════════════════════
# Setup decisions
# ══════════════════════════════════════════════════════════════════════════
r += 2
section_title(ws, r, last_col, "Setup decisions")
r += 1
ws.row_dimensions[r].height = 16
hdr(ws.cell(r, 2), "Part")
for t in periods:
    hdr(ws.cell(r, t + 2), str(t))

for idx, i in enumerate(PARTS, start=1):
    r += 1
    ws.row_dimensions[r].height = 15
    plain(ws.cell(r, 2), i, size=9, border=bot_thin)
    for t in periods:
        setup = 1 if schedule[i][t] > 0.5 else 0
        plain(ws.cell(r, t + 2), "x" if setup else "",
              bold=True, size=9, align="center", border=bot_thin)

wb.save(OUTPUT_FILE)
print(f"Results written to {OUTPUT_FILE} -> sheet '{SHEET_NAME}'")
print(f"Total cost (1b): EUR {total_cost_1b:,.2f}")
print(f"  Setup:         EUR {total_setup:,.2f}")
print(f"  Holding:       EUR {total_holding:,.2f}")
print(f"  Backorder:     EUR {total_backorder:,.2f}")
print(f"Service level:   {service_level*100:.1f}%")
print(f"Fill rate:       {fill_rate*100:.2f}%")