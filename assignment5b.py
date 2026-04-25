"""
APM Project 2026 - Assignment 5b
"""

import gurobipy as gp
from gurobipy import GRB
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from input_data import (
    PARTS, END_PRODUCT, T, BOM, LEAD_TIME, MIN_LOT,
    INIT_INV, SETUP_COST, HOLDING_COST,
    DEMAND_FORECAST, DEMAND_REALIZED, BACKORDER_COST
)

def get_parents(part):
    parents = {}
    for parent, children in BOM.items():
        if part in children:
            parents[parent] = children[part]
    return parents

def clean_num(val, tol=1e-6):
    v = float(val)
    return 0.0 if abs(v) < tol else round(v, 6)

# ── Re-solve 5a to get fixed plan ─────────────────────────────────────────
CAP_X_BASE     = 800
CAP_X_MAX_EXP  = 200
COST_EXP_X     = 10
CAP_X_OT_MAX   = 300
COST_OT_X      = 2
CAP_Y_BASE     = 60 * 24 * 7 - 80
CAP_Y_MAX_PCT  = 40
COST_EXP_Y_PCT = 1_500
CAP_Y_OT_MAX   = 38
COST_OT_Y      = 120
PROC_X = {END_PRODUCT: 1}
PROC_Y = {"B1401": 3, "B2302": 2}

periods = range(1, T + 1)

m5a = gp.Model("APM_5a_reload")
m5a.setParam("OutputFlag", 0)
xa     = m5a.addVars(PARTS, periods, name="x",  vtype=GRB.INTEGER, lb=0)
ya     = m5a.addVars(PARTS, periods, name="y",  vtype=GRB.BINARY)
Ia     = m5a.addVars(PARTS, periods, name="I",  vtype=GRB.INTEGER, lb=0)
dxa    = m5a.addVar(name="dx",     vtype=GRB.INTEGER, lb=0, ub=CAP_X_MAX_EXP)
dya    = m5a.addVar(name="dy_pct", lb=0, ub=CAP_Y_MAX_PCT)
ot_xa  = m5a.addVars(periods, name="ot_x", vtype=GRB.INTEGER, lb=0, ub=CAP_X_OT_MAX)
ot_ya  = m5a.addVars(periods, name="ot_y", lb=0, ub=CAP_Y_OT_MAX)
BIG_M  = {i: sum(DEMAND_FORECAST)*25 for i in PARTS}

m5a.setObjective(
    gp.quicksum(SETUP_COST[i]*ya[i,t] + HOLDING_COST[i]*Ia[i,t]
                for i in PARTS for t in periods)
    + COST_EXP_X*dxa + COST_EXP_Y_PCT*dya
    + gp.quicksum(COST_OT_X*ot_xa[t] + COST_OT_Y*ot_ya[t] for t in periods),
    GRB.MINIMIZE)

for i in PARTS:
    for t in periods:
        inv_prev = INIT_INV[i] if t==1 else Ia[i,t-1]
        op = t - LEAD_TIME[i]
        rec = xa[i,op] if op>=1 else 0
        int_d = gp.quicksum(BOM[p][i]*xa[p,t] for p in get_parents(i))
        ext_d = DEMAND_FORECAST[t-1] if i==END_PRODUCT else 0
        m5a.addConstr(Ia[i,t] == inv_prev + rec - int_d - ext_d)
        m5a.addConstr(xa[i,t] >= MIN_LOT[i]*ya[i,t])
        m5a.addConstr(xa[i,t] <= BIG_M[i]*ya[i,t])

for t in periods:
    m5a.addConstr(gp.quicksum(PROC_X.get(i,0)*xa[i,t] for i in PARTS)
                  <= CAP_X_BASE + dxa + ot_xa[t])
    m5a.addConstr(gp.quicksum(PROC_Y.get(i,0)*xa[i,t] for i in PARTS)
                  <= CAP_Y_BASE + (CAP_Y_BASE/100)*dya + 60*ot_ya[t])
m5a.optimize()

dx_val     = clean_num(dxa.X)
dy_pct_val = clean_num(dya.X)
invest_X   = COST_EXP_X     * dx_val
invest_Y   = COST_EXP_Y_PCT * dy_pct_val
cap_x_new  = CAP_X_BASE + dx_val
cap_y_new  = CAP_Y_BASE * (1 + dy_pct_val/100)
cost_5a    = m5a.ObjVal
total_ot_x_5a = sum(COST_OT_X * ot_xa[t].X for t in periods)
total_ot_y_5a = sum(COST_OT_Y * ot_ya[t].X for t in periods)
ot_x_vals  = {t: clean_num(ot_xa[t].X) for t in periods}
ot_y_vals  = {t: clean_num(ot_ya[t].X) for t in periods}

schedule = {i: {t: xa[i,t].X for t in periods} for i in PARTS}
setup_cost_fixed = sum(SETUP_COST[i] for i in PARTS for t in periods if xa[i,t].X > 0.5)

# ── Simulate with realized demand ─────────────────────────────────────────
inventory = {i: {} for i in PARTS}
held      = {i: {} for i in PARTS}
backorders = {t: 0.0 for t in periods}
net_end    = {}

for i in PARTS:
    if i == END_PRODUCT:
        continue
    inv_prev = float(INIT_INV[i])
    for t in periods:
        op = t - LEAD_TIME[i]
        rec = schedule[i].get(op, 0.0) if op >= 1 else 0.0
        int_d = sum(BOM[p][i]*schedule[p].get(t,0.0) for p in get_parents(i))
        net = inv_prev + rec - int_d
        inventory[i][t] = clean_num(net)
        held[i][t]      = clean_num(max(0.0, net))
        inv_prev = net

inv_prev = float(INIT_INV[END_PRODUCT])
bo_prev  = 0.0
for t in periods:
    op = t - LEAD_TIME[END_PRODUCT]
    rec = schedule[END_PRODUCT].get(op, 0.0) if op >= 1 else 0.0
    net = inv_prev - bo_prev + rec - DEMAND_REALIZED[t-1]
    net_end[t] = clean_num(net)
    if net >= 0:
        inventory[END_PRODUCT][t] = clean_num(net)
        held[END_PRODUCT][t]      = clean_num(net)
        backorders[t] = 0.0
    else:
        inventory[END_PRODUCT][t] = 0.0
        held[END_PRODUCT][t]      = 0.0
        backorders[t] = clean_num(-net)
    inv_prev = inventory[END_PRODUCT][t]
    bo_prev  = backorders[t]

total_holding   = sum(HOLDING_COST[i]*held[i][t] for i in PARTS for t in periods)
total_backorder = sum(BACKORDER_COST * backorders[t] for t in periods)
total_cost_5b   = setup_cost_fixed + total_holding + total_backorder + invest_X + invest_Y + total_ot_x_5a + total_ot_y_5a

periods_no_bo = sum(1 for t in periods if backorders[t] == 0)
service_level = periods_no_bo / T
total_demand  = sum(DEMAND_REALIZED)
total_new_bo  = 0.0
prev_bo = 0.0
for t in periods:
    new_bo = max(0.0, backorders[t] - prev_bo)
    total_new_bo += new_bo
    prev_bo = backorders[t]
fill_rate     = 1.0 - (total_new_bo / total_demand) if total_demand > 0 else 1.0
units_on_time = total_demand - total_new_bo

# ── Read reference results from OUTPUT.xlsx ───────────────────────────────
def read_float(ws, row, col=3):
    try: return float(ws.cell(row, col).value)
    except: return None

import re
def parse_pct(s):
    m = re.search(r"[\d.]+", str(s) if s else "")
    return float(m.group()) if m else None

refs = {}
try:
    wb_ref = load_workbook("OUTPUT.xlsx", data_only=True)
    ws1 = wb_ref["Output_1b"]
    refs["1b"] = dict(setup=read_float(ws1,3), holding=read_float(ws1,4),
                      bo=read_float(ws1,5),    total=read_float(ws1,6),
                      sl=parse_pct(ws1.cell(11,3).value), fr=parse_pct(ws1.cell(12,3).value))
    ws2 = wb_ref["Output_2b"]
    refs["2b"] = dict(setup=read_float(ws2,3), holding=read_float(ws2,4),
                      bo=read_float(ws2,5),    total=read_float(ws2,6),
                      sl=parse_pct(ws2.cell(9,3).value),  fr=parse_pct(ws2.cell(10,3).value))
    ws3 = wb_ref["Output_3b"]
    refs["3b"] = dict(setup=read_float(ws3,3), holding=read_float(ws3,4),
                      bo=read_float(ws3,7),    total=read_float(ws3,8),
                      sl=parse_pct(ws3.cell(13,3).value), fr=parse_pct(ws3.cell(14,3).value))
    ws4 = wb_ref["Output_4b"]
    refs["4b"] = dict(setup=read_float(ws4,3), holding=read_float(ws4,4),
                      bo=read_float(ws4,5),    total=read_float(ws4,6),
                      sl=parse_pct(ws4.cell(13,3).value), fr=parse_pct(ws4.cell(14,3).value))
    has_ref = True
except Exception:
    has_ref = False

# ── Write to OUTPUT.xlsx ───────────────────────────────────────────────────
OUTPUT_FILE = "OUTPUT.xlsx"
SHEET_NAME  = "Output_5b"

if os.path.exists(OUTPUT_FILE):
    wb = load_workbook(OUTPUT_FILE)
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)
else:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

NO_FILL    = PatternFill(fill_type=None)
BLACK_FILL = PatternFill("solid", start_color="000000", end_color="000000")
none_border = Border()
bot_medium  = Border(bottom=Side(style="medium", color="000000"))
bot_thin    = Border(bottom=Side(style="thin",   color="CCCCCC"))

def plain(cell, val, bold=False, align="left", size=9, color="000000", fmt=None, border=None):
    cell.value = val
    cell.font  = Font(name="Calibri", bold=bold, size=size, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.fill  = NO_FILL
    if border: cell.border = border
    if fmt:    cell.number_format = fmt

def hdr(cell, val):
    cell.value = val
    cell.font  = Font(name="Calibri", bold=False, size=9, color="FFFFFF")
    cell.fill  = BLACK_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = none_border

def section_title(ws, r, last_col, text):
    ws.row_dimensions[r].height = 14
    plain(ws.cell(r, 2), text, size=8, color="888888", border=bot_medium)
    for col in range(3, last_col + 1):
        ws.cell(r, col).border = bot_medium

ws.column_dimensions["A"].width = 1
ws.column_dimensions["B"].width = 9
for col in range(3, T + 4):
    ws.column_dimensions[get_column_letter(col)].width = 5.5
last_col = T + 2

# ── Title ──────────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 26
ws.merge_cells(f"B1:{get_column_letter(last_col)}1")
plain(ws.cell(1, 2), "Assignment 5b - Realized demand evaluation (overtime + permanent expansion)", bold=True, size=13)

# ── Cost summary ───────────────────────────────────────────────────────────
ws.row_dimensions[2].height = 4
cost_rows = [
    ("Setup cost",          setup_cost_fixed,  False),
    ("Holding cost",        total_holding,     False),
    ("Backorder cost",      total_backorder,   False),
    ("Investment cost X",   invest_X,          False),
    ("Investment cost Y",   invest_Y,          False),
    ("Overtime cost X",     total_ot_x_5a,     False),
    ("Overtime cost Y",     total_ot_y_5a,     False),
    ("Total cost (5b)",     total_cost_5b,     True),
    ("Total cost (5a)",     cost_5a,           False),
    ("Difference",          total_cost_5b - cost_5a, False),
]
for r, (label, val, bold) in enumerate(cost_rows, start=3):
    ws.row_dimensions[r].height = 16
    plain(ws.cell(r, 2), label, size=9, color="555555")
    fmt = '"+"€"#,##0.00;"-"€"#,##0.00' if label == "Difference" else '"€"#,##0.00'
    plain(ws.cell(r, 3), round(clean_num(val), 2), bold=bold, size=9, fmt=fmt)
    ws.merge_cells(f"C{r}:{get_column_letter(last_col)}{r}")

# ── Service metrics ────────────────────────────────────────────────────────
ws.row_dimensions[13].height = 6
ws.row_dimensions[14].height = 14
section_title(ws, 14, last_col, "Service metrics (end product)")
for r, (label, text, red) in enumerate([
    ("Service level",   f"{service_level*100:.1f}%  ({periods_no_bo}/{T} periods without backorder)", False),
    ("Fill rate",       f"{fill_rate*100:.2f}%  ({units_on_time:,.0f} / {total_demand:,.0f} units on time)", False),
    ("Total backorder", f"{total_new_bo:,.0f} units", total_new_bo > 0),
], start=15):
    ws.row_dimensions[r].height = 16
    plain(ws.cell(r, 2), label, size=9, color="555555")
    plain(ws.cell(r, 3), text, size=9, color="CC0000" if red else "000000")
    ws.merge_cells(f"C{r}:{get_column_letter(last_col)}{r}")

# ══════════════════════════════════════════════════════════════════════════
# Comparison table 1b / 2b / 3b / 4b / 5b
# ══════════════════════════════════════════════════════════════════════════
r = 19
section_title(ws, r, last_col, "Comparison: 1b vs 2b vs 3b vs 4b vs 5b")
r += 1
ws.row_dimensions[r].height = 16

for c, w in [(2,22),(3,13),(4,13),(5,13),(6,13),(7,13)]:
    ws.column_dimensions[get_column_letter(c)].width = w

hdr(ws.cell(r, 2), "Metric")
for col, lbl in [(3,"1b"),(4,"2b"),(5,"3b"),(6,"4b"),(7,"5b")]:
    hdr(ws.cell(r, col), lbl)

cmp_rows = [
    ("Total cost (EUR)",    "total",   '"€"#,##0.00'),
    ("Setup cost (EUR)",    "setup",   '"€"#,##0.00'),
    ("Holding cost (EUR)",  "holding", '"€"#,##0.00'),
    ("Backorder cost (EUR)","bo",      '"€"#,##0.00'),
    ("Service level (%)",   "sl",      '0.0'),
    ("Fill rate (%)",       "fr",      '0.00'),
]
vals_5b = dict(total=total_cost_5b, setup=setup_cost_fixed,
               holding=total_holding, bo=total_backorder,
               sl=service_level*100, fr=fill_rate*100)

for label, key, fmt in cmp_rows:
    r += 1
    ws.row_dimensions[r].height = 15
    plain(ws.cell(r, 2), label, size=9, border=bot_thin)
    for col, src in [(3,"1b"),(4,"2b"),(5,"3b"),(6,"4b")]:
        v = refs.get(src, {}).get(key) if has_ref else None
        if v is not None:
            plain(ws.cell(r, col), round(float(v), 2), size=9,
                  align="right", fmt=fmt, border=bot_thin)
        else:
            plain(ws.cell(r, col), "n/a", size=9, align="right",
                  color="AAAAAA", border=bot_thin)
    plain(ws.cell(r, 7), round(vals_5b[key], 2), bold=True,
          size=9, align="right", fmt=fmt, border=bot_thin)

# reset week column widths
ws.column_dimensions["B"].width = 9
for col in range(3, T + 4):
    ws.column_dimensions[get_column_letter(col)].width = 5.5

# ══════════════════════════════════════════════════════════════════════════
# Backorder schedule
# ══════════════════════════════════════════════════════════════════════════
r += 2
section_title(ws, r, last_col, f"Backorder schedule - {END_PRODUCT} (end product only)")
r += 1
ws.row_dimensions[r].height = 16
hdr(ws.cell(r, 2), "Part")
for t in periods:
    hdr(ws.cell(r, t+2), str(t))

r += 1
ws.row_dimensions[r].height = 15
plain(ws.cell(r, 2), END_PRODUCT, size=9, border=bot_thin)
for t in periods:
    v = round(backorders[t])
    if v > 0:
        plain(ws.cell(r, t+2), v, bold=True, size=9, align="center",
              color="CC0000", fmt='#,##0', border=bot_thin)
    else:
        plain(ws.cell(r, t+2), "", size=9, align="center", border=bot_thin)

# ══════════════════════════════════════════════════════════════════════════
# Demand vs delivered
# ══════════════════════════════════════════════════════════════════════════
r += 2
section_title(ws, r, last_col, "Demand vs delivered (end product)")
r += 1
ws.row_dimensions[r].height = 16
hdr(ws.cell(r, 2), "")
for t in periods:
    hdr(ws.cell(r, t+2), str(t))

r += 1
ws.row_dimensions[r].height = 15
plain(ws.cell(r, 2), "Demand", size=9, border=bot_thin)
for t in periods:
    plain(ws.cell(r, t+2), DEMAND_REALIZED[t-1], size=9, align="center",
          fmt='#,##0', border=bot_thin)

r += 1
ws.row_dimensions[r].height = 15
plain(ws.cell(r, 2), "Delivered", size=9, border=bot_thin)
prev_bo = 0.0
for t in periods:
    new_bo_t = max(0.0, backorders[t] - prev_bo)
    delivered = DEMAND_REALIZED[t-1] - new_bo_t
    shortage  = delivered < DEMAND_REALIZED[t-1]
    plain(ws.cell(r, t+2), round(delivered), size=9, align="center",
          color="CC0000" if shortage else "000000",
          fmt='#,##0', border=bot_thin)
    prev_bo = backorders[t]

# ══════════════════════════════════════════════════════════════════════════
# Production schedule
# ══════════════════════════════════════════════════════════════════════════
r += 2
section_title(ws, r, last_col, "Production / order schedule (units, from 5a plan)")
r += 1
ws.row_dimensions[r].height = 16
hdr(ws.cell(r, 2), "Part")
for t in periods:
    hdr(ws.cell(r, t+2), str(t))

for i in PARTS:
    r += 1
    ws.row_dimensions[r].height = 15
    plain(ws.cell(r, 2), i, size=9, border=bot_thin)
    for t in periods:
        val = round(schedule[i][t]) if schedule[i][t] > 0.5 else ""
        plain(ws.cell(r, t+2), val, bold=bool(val), size=9, align="center",
              fmt='#,##0' if val != "" else None, border=bot_thin)

# ══════════════════════════════════════════════════════════════════════════
# Inventory levels (realized)
# ══════════════════════════════════════════════════════════════════════════
r += 2
section_title(ws, r, last_col, "Inventory levels (end of period, realized)")
r += 1
ws.row_dimensions[r].height = 16
hdr(ws.cell(r, 2), "Part")
for t in periods:
    hdr(ws.cell(r, t+2), str(t))

for i in PARTS:
    r += 1
    ws.row_dimensions[r].height = 15
    plain(ws.cell(r, 2), i, size=9, border=bot_thin)
    for t in periods:
        v = round(inventory[i][t])
        plain(ws.cell(r, t+2), v, size=9, align="center",
              color="BBBBBB" if v == 0 else "000000",
              fmt='#,##0', border=bot_thin)

# ══════════════════════════════════════════════════════════════════════════
# Overtime usage (from 5a plan)
# ══════════════════════════════════════════════════════════════════════════
r += 2
section_title(ws, r, last_col, "Overtime usage (from 5a plan, unchanged)")
r += 1
ws.row_dimensions[r].height = 16
hdr(ws.cell(r, 2), "")
for t in periods:
    hdr(ws.cell(r, t+2), str(t))

for label, vals, fmt in [
    ("X  (units OT)", ot_x_vals, '#,##0'),
    ("Y  (hours OT)", ot_y_vals, '0.00'),
]:
    r += 1
    ws.row_dimensions[r].height = 15
    plain(ws.cell(r, 2), label, size=9, border=bot_thin)
    for t in periods:
        v = round(vals[t], 2)
        col = "CC0000" if v else "BBBBBB"
        plain(ws.cell(r, t+2), v if v else "", bold=bool(v), size=9,
              align="center", color=col,
              fmt=fmt if v else None, border=bot_thin)

# ══════════════════════════════════════════════════════════════════════════
# Setup decisions
# ══════════════════════════════════════════════════════════════════════════
r += 2
section_title(ws, r, last_col, "Setup decisions")
r += 1
ws.row_dimensions[r].height = 16
hdr(ws.cell(r, 2), "Part")
for t in periods:
    hdr(ws.cell(r, t+2), str(t))

for i in PARTS:
    r += 1
    ws.row_dimensions[r].height = 15
    plain(ws.cell(r, 2), i, size=9, border=bot_thin)
    for t in periods:
        plain(ws.cell(r, t+2), "x" if schedule[i][t] > 0.5 else "",
              bold=True, size=9, align="center", border=bot_thin)

wb.save(OUTPUT_FILE)
print(f"Results written to {OUTPUT_FILE} -> sheet '{SHEET_NAME}'")
print(f"Total cost (5b): EUR {clean_num(total_cost_5b):,.2f}")
print(f"  Setup:         EUR {clean_num(setup_cost_fixed):,.2f}")
print(f"  Holding:       EUR {clean_num(total_holding):,.2f}")
print(f"  Backorder:     EUR {clean_num(total_backorder):,.2f}")
print(f"  Investment X:  EUR {clean_num(invest_X):,.2f}")
print(f"  Investment Y:  EUR {clean_num(invest_Y):,.2f}")
print(f"  Overtime X:    EUR {clean_num(total_ot_x_5a):,.2f}")
print(f"  Overtime Y:    EUR {clean_num(total_ot_y_5a):,.2f}")
print(f"Service level:   {service_level*100:.1f}%")
print(f"Fill rate:       {fill_rate*100:.2f}%")